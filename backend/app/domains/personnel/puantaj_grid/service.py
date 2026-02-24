"""Puantaj Grid domain service"""
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
from datetime import date, datetime
from itertools import groupby
import calendar

from app.domains.personnel.puantaj_grid.repository import PuantajGridRepository
from app.models import PersonnelPuantajGrid, PuantajDurum


class PuantajGridService:
    """Service layer for Puantaj Grid"""
    
    def __init__(self, db: Session):
        self.db = db
        self.repository = PuantajGridRepository()
    
    def get_grid_data(
        self,
        donem: str,
        cost_center_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Get grid data for a period
        Returns Excel-like grid with personnel and daily status
        """
        # Parse period
        year, month = map(int, donem.split('-'))
        donem_ilk_gun = date(year, month, 1)
        son_gun = calendar.monthrange(year, month)[1]
        donem_son_gun = date(year, month, son_gun)
        
        # Get holidays from calendar_holidays table
        holiday_days = set()
        try:
            from sqlalchemy import text
            result = self.db.execute(
                text("SELECT DAY(holiday_date) as gun FROM calendar_holidays WHERE YEAR(holiday_date) = :year AND MONTH(holiday_date) = :month"),
                {"year": year, "month": month}
            )
            holiday_days = {row.gun for row in result}
        except Exception as e:
            print(f"[WARNING] Could not fetch holidays: {e}")
            holiday_days = set()
        
        # Get grid records
        grid_records = self.repository.get_grid_records(
            self.db,
            donem,
            cost_center_id
        )
        
        result = []
        
        # DEBUG: Grid kayıtlarını kontrol et
        if grid_records:
            print(f"[DEBUG] Grid records found: {len(grid_records)} unique personnel")
        
        if grid_records:
            # Grid records exist - process them
            personnel_ids = [g.personnel_id for g in grid_records]
            
            # Batch fetch personnel
            personnel_dict = self.repository.get_personnel_batch(self.db, personnel_ids)
            
            # Process grid records - her grid için SADECE bir satır
            for grid in grid_records:
                person = personnel_dict.get(grid.personnel_id)
                
                if person:
                    # Draft contract - personnel_id bazlı sorgulama
                    from app.models import PersonnelDraftContract
                    draft = self.db.query(PersonnelDraftContract).filter(
                        PersonnelDraftContract.personnel_id == person.id,
                        PersonnelDraftContract.is_active == 1
                    ).first()
                    
                    # Draft contract yoksa personeli gösterme
                    if not draft:
                        continue
                    
                    # Maliyet merkezi filtresi - DRAFT CONTRACT'TAKİ güncel değeri kullan
                    if cost_center_id and draft.cost_center_id != cost_center_id:
                        continue
                    
                    # Actual contract bilgisi için personnel_id ile son sözleşmeyi al (aktif/pasif farketmez)
                    from app.models import PersonnelContract
                    contract = self.db.query(PersonnelContract).filter(
                        PersonnelContract.personnel_id == person.id
                    ).order_by(PersonnelContract.ise_giris_tarihi.desc()).first()
                    
                    row = {
                        'id': person.id,
                        'adi_soyadi': f"{person.ad} {person.soyad}",
                            'tc_kimlik_no': person.tc_kimlik_no or '',
                            'cost_center_id': contract.cost_center_id if contract else None,
                            'taseron_name': getattr(contract, '_taseron_name', None) if contract else None,
                            'meslek_adi': getattr(contract, '_meslek_adi', None) if contract else None,
                            'draft_contract_id': draft.id if draft else None,
                            'calisma_takvimi': draft.calisma_takvimi if draft else None,
                            'ucret_nevi': draft.ucret_nevi if draft else None,
                            'fm_orani': float(draft.fm_orani) if draft else None,
                            'tatil_orani': float(draft.tatil_orani) if draft else None,
                            'maas1_tip': contract.net_brut if contract else None,
                            'maas1_tutar': float(contract.ucret) if (contract and contract.ucret) else None,
                            'maas2_tutar': float(draft.net_ucret) if draft else None
                        }
                else:
                    row = {
                        'id': grid.personnel_id,
                        'adi_soyadi': f'Personel {grid.personnel_id}',
                        'tc_kimlik_no': '',
                        'cost_center_id': None,
                        'taseron_name': None,
                        'meslek_adi': None,
                        'draft_contract_id': None,
                        'calisma_takvimi': None,
                        'ucret_nevi': None,
                        'fm_orani': None,
                        'tatil_orani': None,
                        'net_brut': None,
                        'ucret': None,
                        'maas2_tutar': None
                    }
                
                # Personelin çalışma tarihlerini al
                ise_giris = contract.ise_giris_tarihi if contract else None
                isten_cikis = contract.isten_cikis_tarihi if contract else None
                
                row['ise_giris_tarihi'] = ise_giris.isoformat() if ise_giris else None
                row['isten_cikis_tarihi'] = isten_cikis.isoformat() if isten_cikis else None
                
                # Calisma takvimi - draft contract'tan al
                draft = contract.draft_contract if contract else None
                calisma_takvimi = draft.calisma_takvimi if draft else None
                
                # Add 31 daily columns
                for i in range(1, 32):
                    gun_col = f'gun_{i}'
                    fm_col = f'fm_gun_{i}'
                    
                    val = getattr(grid, gun_col, None)
                    
                    # Sigortasının olmadığı günleri kontrol et
                    is_not_insured = False
                    if i <= son_gun:
                        current_date = date(year, month, i)
                        if ise_giris and current_date < ise_giris:
                            is_not_insured = True
                        elif isten_cikis and current_date > isten_cikis:
                            is_not_insured = True
                    
                    # Sigortası yoksa VE kullanıcı değer girmemişse '-' işareti koy
                    if is_not_insured and (val is None or val == ''):
                        row[gun_col] = '-'
                    # Değer yoksa ve tatil günüyse 'T'
                    elif (val is None or val == '') and i in holiday_days:
                        row[gun_col] = 'T'
                    # Değer yoksa ve pazar günüyse ve atipi ise 'H'
                    elif (val is None or val == '') and i <= son_gun:
                        current_date = date(year, month, i)
                        is_sunday = current_date.weekday() == 6
                        if is_sunday and calisma_takvimi == 'atipi':
                            row[gun_col] = 'H'
                        else:
                            row[gun_col] = val.value if hasattr(val, 'value') else val
                    else:
                        # Kullanıcı değer girmiş - koru (sigortasız bile olsa)
                        row[gun_col] = val.value if hasattr(val, 'value') else val
                    
                    fm_val = getattr(grid, fm_col, None)
                    row[fm_col] = float(fm_val) if fm_val is not None else None
                
                # Add department for sorting
                row['_departman'] = contract.departman if contract and contract.departman else 'Departman Belirtilmemiş'
                
                result.append(row)
        else:
            # No grid records - fetch from personnel/contracts
            personnel_ids = self.repository.get_active_personnel_ids(
                self.db,
                donem_ilk_gun,
                donem_son_gun,
                cost_center_id
            )
            
            if personnel_ids:
                # Batch fetch personnel and contracts
                personnel_dict = self.repository.get_personnel_batch(self.db, personnel_ids)
                contract_dict = self.repository.get_contracts_batch(
                    self.db,
                    personnel_ids,
                    donem_ilk_gun,
                    donem_son_gun
                )
                
                for person_id in personnel_ids:
                    person = personnel_dict.get(person_id)
                    contracts = contract_dict.get(person_id, [])
                    
                    if not person:
                        continue
                    
                    # Draft contract kontrolü - yoksa personeli gösterme
                    from app.models import PersonnelDraftContract
                    draft = self.db.query(PersonnelDraftContract).filter(
                        PersonnelDraftContract.personnel_id == person.id,
                        PersonnelDraftContract.is_active == 1
                    ).first()
                    
                    if not draft:
                        continue  # Draft contract yoksa bu personeli gösterme
                    
                    # DEBUG - Selman için özel kontrol
                    if "SELMAN" in person.ad.upper() and "COŞKUN" in person.soyad.upper():
                        print(f"🔍 SELMAN BULUNDU!")
                        print(f"  - draft.cost_center_id: {draft.cost_center_id}")
                        print(f"  - filter cost_center_id: {cost_center_id}")
                        print(f"  - Eşleşme: {draft.cost_center_id == cost_center_id}")
                    
                    # Maliyet merkezi filtresi - DRAFT CONTRACT'TAKİ güncel değeri kullan
                    if cost_center_id and draft.cost_center_id != cost_center_id:
                        if "SELMAN" in person.ad.upper() and "COŞKUN" in person.soyad.upper():
                            print(f"  ❌ FİLTRELENDİ!")
                        continue
                    
                    if "SELMAN" in person.ad.upper() and "COŞKUN" in person.soyad.upper():
                        print(f"  ✅ LİSTEYE EKLENDİ!")
                    
                    # Contract bilgilerini al (sadece bilgi amaçlı, filtreleme için değil)
                    contracts = contract_dict.get(person_id, [])
                    contract = contracts[0] if contracts else None
                    
                    adi_soyadi = f"{person.ad} {person.soyad}"
                    
                    row = {
                        'id': person.id,
                        'adi_soyadi': adi_soyadi,
                        'tc_kimlik_no': person.tc_kimlik_no or '',
                        'cost_center_id': contract.cost_center_id if contract else None,
                        'taseron_name': getattr(contract, '_taseron_name', None) if contract else None,
                        'meslek_adi': getattr(contract, '_meslek_adi', None) if contract else None,
                        'draft_contract_id': draft.id if draft else None,
                        'calisma_takvimi': draft.calisma_takvimi if draft else None,
                        'ucret_nevi': draft.ucret_nevi if draft else None,
                        'fm_orani': float(draft.fm_orani) if draft else None,
                        'tatil_orani': float(draft.tatil_orani) if draft else None,
                        'net_brut': contract.net_brut if contract else None,
                        'ucret': float(contract.ucret) if (contract and contract.ucret) else None,
                        'maas2_tutar': float(draft.net_ucret) if draft else None
                    }
                    
                    # Personelin çalışma tarihlerini al
                    ise_giris = contract.ise_giris_tarihi if contract else None
                    isten_cikis = contract.isten_cikis_tarihi if contract else None
                    
                    row['ise_giris_tarihi'] = ise_giris.isoformat() if ise_giris else None
                    row['isten_cikis_tarihi'] = isten_cikis.isoformat() if isten_cikis else None
                    
                    # Add 31 daily columns - auto-mark holidays as 'T' and Sundays as 'H' for atipi
                    # calisma_takvimi draft'tan alındı (yukarıda)
                    calisma_takvimi = draft.calisma_takvimi if draft else None
                    for i in range(1, 32):
                        if i <= son_gun:
                            current_date = date(year, month, i)
                            
                            # Sigortasının olmadığı günleri kontrol et
                            is_not_insured = False
                            if ise_giris and current_date < ise_giris:
                                is_not_insured = True
                            elif isten_cikis and current_date > isten_cikis:
                                is_not_insured = True
                            
                            # Sigortası yoksa '-' işareti koy
                            if is_not_insured:
                                row[f'gun_{i}'] = '-'
                            # Tatil günüyse 'T'
                            elif i in holiday_days:
                                row[f'gun_{i}'] = 'T'
                            # Pazar günü mü? (weekday: 0=Pazartesi, 6=Pazar)
                            elif current_date.weekday() == 6 and calisma_takvimi == 'atipi':
                                row[f'gun_{i}'] = 'H'
                            else:
                                row[f'gun_{i}'] = None
                        else:
                            row[f'gun_{i}'] = None
                        row[f'fm_gun_{i}'] = None
                    
                    # Add department for sorting
                    row['_departman'] = contract.departman if contract and contract.departman else 'Departman Belirtilmemiş'
                    
                    result.append(row)
        
        # Sort and group by department
        result = self._sort_and_group_by_department(result)
        
        return {
            "success": True,
            "donem": donem,
            "total": len(result),
            "records": result,
            "holidays": list(holiday_days)
        }
    
    def save_grid_data(
        self,
        donem: str,
        records: List[dict]
    ) -> Dict[str, Any]:
        """
        Save grid data from Excel-like grid
        INSERT/UPDATE 31-day data for each personnel
        """
        year, month = map(int, donem.split('-'))
        saved_count = 0
        updated_count = 0
        
        for record in records:
            personnel_id = record['id']
            cost_center_id = record.get('cost_center_id')  # Frontend'den geliyor
            
            # Check if record exists (bir personelin bir dönemde tek puantajı var)
            existing = self.repository.get_by_personnel_donem(
                self.db,
                personnel_id,
                donem
            )
            
            # Prepare 31-day data
            gun_data = {}
            for i in range(1, 32):
                # Durum kodu (gun_1 - gun_31)
                gun_col = f'gun_{i}'
                val = record.get(gun_col)
                
                # Convert empty values to None, '-' to enum
                # 'MINUS' is used in Excel for days not in the month
                if val in [None, '', 'MINUS', '-']:
                    gun_data[gun_col] = None
                else:
                    # Convert to Enum
                    try:
                        gun_data[gun_col] = PuantajDurum(val)
                    except:
                        gun_data[gun_col] = None
                
                # Fazla mesai (fm_gun_1 - fm_gun_31)
                fm_col = f'fm_gun_{i}'
                fm_val = record.get(fm_col)
                
                # Convert empty or zero values to None
                if fm_val in [None, '', 0, '0']:
                    gun_data[fm_col] = None
                else:
                    try:
                        gun_data[fm_col] = float(fm_val)
                    except:
                        gun_data[fm_col] = None
            
            # Özet alanları ekle (frontend'den hesaplanmış geliyor)
            summary_fields = [
                'calisilan_gun_sayisi', 'ssk_gun_sayisi', 'yillik_izin_gun',
                'izin_gun_sayisi', 'rapor_gun_sayisi', 'eksik_gun_sayisi',
                'yarim_gun_sayisi', 'toplam_gun_sayisi', 'normal_calismasi',
                'fazla_calismasi', 'eksik_calismasi', 'gece_calismasi', 'tatil_calismasi',
                'sigorta_girmedigi', 'hafta_tatili', 'resmi_tatil',
                'yol', 'prim', 'ikramiye', 'bayram', 'kira',
                # Maas2 kazanç hesaplamaları (frontend'den gelir)
                'maas2_gunluk_kazanc', 'maas2_normal_kazanc', 'maas2_mesai_kazanc',
                'maas2_eksik_kazanc', 'maas2_tatil_kazanc', 'maas2_tatil_mesai_kazanc', 'maas2_toplam_kazanc'
            ]
            
            # Decimal/float fields
            decimal_fields = ['yarim_gun_sayisi', 'normal_calismasi', 'fazla_calismasi', 
                             'eksik_calismasi', 'gece_calismasi', 'tatil_calismasi', 'yol', 'prim', 
                             'ikramiye', 'bayram', 'kira',
                             'maas2_gunluk_kazanc', 'maas2_normal_kazanc', 'maas2_mesai_kazanc',
                             'maas2_eksik_kazanc', 'maas2_tatil_kazanc', 'maas2_tatil_mesai_kazanc', 'maas2_toplam_kazanc']
            
            for field in summary_fields:
                val = record.get(field)
                if val is not None and val != '':
                    try:
                        if field in decimal_fields:
                            gun_data[field] = float(val)
                        else:
                            gun_data[field] = int(float(val))  # float'tan int'e çevir (Excel'den 5.0 gelebilir)
                    except:
                        gun_data[field] = 0
            
            if existing:
                # UPDATE
                gun_data['updated_at'] = datetime.now()
                self.repository.update(self.db, existing, **gun_data)
                updated_count += 1
            else:
                # INSERT
                # Calculate days in month
                days_in_month = calendar.monthrange(year, month)[1]
                
                new_record = PersonnelPuantajGrid(
                    personnel_id=personnel_id,
                    cost_center_id=cost_center_id,
                    donem=donem,
                    yil=year,
                    ay=month,
                    ayin_toplam_gun_sayisi=days_in_month,
                    **gun_data
                )
                self.repository.create(self.db, new_record)
                saved_count += 1
        
        self.db.commit()
        
        return {
            "success": True,
            "donem": donem,
            "saved": saved_count,
            "updated": updated_count,
            "total": saved_count + updated_count
        }
    
    def _sort_and_group_by_department(self, result: List[dict]) -> List[dict]:
        """Sort and group results by taseron and department"""
        # Sort function
        def sort_key(row):
            taseron = row.get('taseron_name') or 'Taşeronsuz'
            departman = row.get('_departman', 'Departman Belirtilmemiş')
            
            # Taşeronsuz en üstte
            if taseron == 'Taşeronsuz':
                taseron_sort = '0'
            else:
                taseron_sort = f'1_{taseron}'
            
            # İdare Ekibi departman içinde üstte
            if departman == 'İdare Ekibi':
                departman_sort = '0'
            else:
                departman_sort = f'1_{departman}'
            
            isim = row.get('adi_soyadi', '')
            return (taseron_sort, taseron, departman_sort, departman, isim)
        
        # Sort first
        result.sort(key=sort_key)
        
        # Group by taseron and department, add header rows
        final_result = []
        
        # First group by taseron
        for taseron_name, taseron_group in groupby(result, key=lambda r: r.get('taseron_name') or 'Taşeronsuz'):
            # Taşeron adını kısalt (ilk 2 kelime)
            taseron_display = taseron_name
            if taseron_name != 'Taşeronsuz':
                words = taseron_name.split()
                if len(words) > 2:
                    taseron_display = ' '.join(words[:2])
            
            # Add taseron header
            final_result.append({
                'row_type': 'taseron_header',
                'taseron_name': taseron_name,
                'id': f'taseron_header_{taseron_name}',
                'adi_soyadi': f'🏢 {taseron_display}',
                'tc_kimlik_no': ''
            })
            
            # Convert to list to allow re-iteration
            taseron_list = list(taseron_group)
            
            # Group by department within taseron
            for dept_name, dept_group in groupby(taseron_list, key=lambda r: r.get('_departman', 'Departman Belirtilmemiş')):
                # Add department header
                final_result.append({
                    'row_type': 'header',
                    'departman': dept_name,
                    'taseron_name': taseron_name,
                    'id': f'header_{taseron_name}_{dept_name}',
                    'adi_soyadi': f'  └─ {dept_name}',
                    'tc_kimlik_no': ''
                })
                
                # Add personnel in department
                for person in dept_group:
                    person['row_type'] = 'data'
                    person['departman'] = person.pop('_departman', None)
                    final_result.append(person)
        
        return final_result
    
    def parse_excel_without_saving(self, contents: bytes, donem_or_filename: str) -> Dict[str, Any]:
        """
        Parse Excel file and return data WITHOUT saving to database
        Frontend will merge this with existing grid data
        """
        import openpyxl
        from io import BytesIO
        import re
        
        # Extract donem
        donem = donem_or_filename
        if not re.match(r'^\d{4}-\d{2}$', donem):
            match = re.search(r'(\d{4}-\d{2})', donem)
            if match:
                donem = match.group(1)
            else:
                raise ValueError("Dönem bilgisi bulunamadı.")
        
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Find header row
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if any(cell and ('personel' in str(cell).lower() or 'ad' in str(cell).lower()) for cell in row):
                header_row = i
                break
        
        if not header_row:
            raise ValueError("Excel başlık satırı bulunamadı")
        
        headers = [cell.value if cell.value else f"col_{i}" for i, cell in enumerate(ws[header_row], start=1)]
        headers_lower = [h.lower().strip() if isinstance(h, str) else h for h in headers]
        
        # Find columns
        tc_col = None
        for i, h in enumerate(headers_lower):
            if h and ('tc' in h or 'kimlik' in h):
                tc_col = i
                break
        
        if tc_col is None:
            raise ValueError("TC Kimlik No kolonu bulunamadı")
        
        # Find day columns
        day_columns = {}
        fm_columns = {}
        summary_columns = {}  # Hesaplama kolonları için
        
        for i, h in enumerate(headers_lower):
            if not h:
                continue
            h_str = str(h).strip()
            
            # gun_1, gun_2, ... or 1, 2, 3, ...
            match = re.match(r'(?:gun[_\s]*)?(\d+)$', h_str)
            if match and 'fm' not in h_str:
                day_num = int(match.group(1))
                if 1 <= day_num <= 31:
                    day_columns[day_num] = i
            # fm_gun_1, fm_1, FM_1, ...
            elif 'fm' in h_str:
                match_fm = re.match(r'(?:fm[_\s]*(?:gun[_\s]*)?)?(\d+)$', h_str)
                if match_fm:
                    day_num = int(match_fm.group(1))
                    if 1 <= day_num <= 31:
                        fm_columns[day_num] = i
            
            # Summary/calculation columns
            if 'çalışılan' in h_str or 'calisilan' in h_str:
                summary_columns['calisilan_gun_sayisi'] = i
            elif 'ssk' in h_str and 'gun' in h_str:
                summary_columns['ssk_gun_sayisi'] = i
            elif 'yıllık' in h_str or 'yillik' in h_str:
                summary_columns['yillik_izin_gun'] = i
            elif 'izin' in h_str and 'gun' in h_str and 'yıllık' not in h_str:
                summary_columns['izin_gun_sayisi'] = i
            elif 'rapor' in h_str:
                summary_columns['rapor_gun_sayisi'] = i
            elif 'eksik' in h_str:
                summary_columns['eksik_gun_sayisi'] = i
            elif 'yarım' in h_str or 'yarim' in h_str:
                summary_columns['yarim_gun_sayisi'] = i
            elif 'toplam' in h_str and 'gun' in h_str:
                summary_columns['toplam_gun_sayisi'] = i
            elif 'normal' in h_str and 'çalışma' in h_str:
                summary_columns['normal_calismasi'] = i
            elif 'fazla' in h_str and 'çalışma' in h_str:
                summary_columns['fazla_calismasi'] = i
            elif 'gece' in h_str and 'çalışma' in h_str:
                summary_columns['gece_calismasi'] = i
            elif 'tatil' in h_str and 'çalışma' in h_str:
                summary_columns['tatil_calismasi'] = i
            elif 'sigorta' in h_str and 'girmedigi' in h_str:
                summary_columns['sigorta_girmedigi'] = i
            elif 'hafta' in h_str and 'tatil' in h_str:
                summary_columns['hafta_tatili'] = i
            elif 'resmi' in h_str and 'tatil' in h_str:
                summary_columns['resmi_tatil'] = i
            elif h_str == 'yol':
                summary_columns['yol'] = i
            elif h_str == 'prim':
                summary_columns['prim'] = i
            elif h_str == 'ikramiye':
                summary_columns['ikramiye'] = i
            elif h_str == 'bayram':
                summary_columns['bayram'] = i
            elif h_str == 'kira':
                summary_columns['kira'] = i
                if 1 <= day_num <= 31:
                    fm_columns[day_num] = i
        
        # Process rows
        records = []
        from app.models import Personnel
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            row_values = [cell.value for cell in row]
            
            if not any(row_values):
                continue
            
            tc_kimlik = row_values[tc_col] if tc_col < len(row_values) else None
            if not tc_kimlik:
                continue
            
            tc_kimlik = str(tc_kimlik).strip()
            
            # Find personnel
            personnel = self.db.query(Personnel).filter(
                Personnel.tc_kimlik_no == tc_kimlik
            ).first()
            
            if not personnel:
                continue
            
            # Build record
            record = {
                'id': personnel.id,
                'tckn': tc_kimlik,  # Frontend 'tckn' alanını kullanıyor
                'tc_kimlik_no': tc_kimlik,
                'sicil_no': tc_kimlik,  # Personnel'de sicil_no yok, tc_kimlik_no kullan
                'adi_soyadi': f"{personnel.ad} {personnel.soyad}"
            }
            
            # Add daily data
            for day_num in range(1, 32):
                gun_key = f'gun_{day_num}'
                fm_key = f'fm_gun_{day_num}'
                
                if day_num in day_columns:
                    val = row_values[day_columns[day_num]] if day_columns[day_num] < len(row_values) else None
                    record[gun_key] = str(val).strip().upper() if val and isinstance(val, str) else (str(val) if val else None)
                else:
                    record[gun_key] = None
                
                if day_num in fm_columns:
                    fm_val = row_values[fm_columns[day_num]] if fm_columns[day_num] < len(row_values) else None
                    try:
                        record[fm_key] = float(fm_val) if fm_val else None
                    except:
                        record[fm_key] = None
                else:
                    record[fm_key] = None
            
            # Excel'den sadece gun_X ve fm_gun_X yüklenir
            # Hesaplanan alanlar (calisilan_gun_sayisi, vb.) detay modalında elle girilir/hesaplanır
            # Bu sayede kullanıcı Excel'den sadece temel verileri yükler, hesaplamaları sonra yapar
            
            records.append(record)
        
        return {
            "success": True,
            "donem": donem,
            "records": records,
            "total": len(records)
        }
    
    def upload_from_excel(self, contents: bytes, donem_or_filename: str) -> Dict[str, Any]:
        """
        Upload puantaj grid data from Excel file
        
        Args:
            contents: Excel file contents
            donem_or_filename: Period (YYYY-MM) or filename containing period
            
        Returns:
            {
                success: true,
                saved: number of new records,
                updated: number of updated records,
                total: total processed
            }
        """
        import openpyxl
        from io import BytesIO
        import re
        
        # Extract donem from filename or use directly
        donem = donem_or_filename
        if not re.match(r'^\d{4}-\d{2}$', donem):
            # Try to extract from filename like "puantaj_2024-01.xlsx"
            match = re.search(r'(\d{4}-\d{2})', donem)
            if match:
                donem = match.group(1)
            else:
                raise ValueError("Dönem bilgisi bulunamadı. Dosya adında YYYY-MM formatında dönem belirtin veya donem parametresini kullanın.")
        
        year, month = map(int, donem.split('-'))
        
        # Load workbook
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Find header row (first row with "Personel" or "Ad Soyad")
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if any(cell and ('personel' in str(cell).lower() or 'ad' in str(cell).lower() or 'soyad' in str(cell).lower()) for cell in row):
                header_row = i
                break
        
        if not header_row:
            raise ValueError("Excel başlık satırı bulunamadı. İlk satırda 'Personel' veya 'Ad Soyad' içeren kolon olmalı.")
        
        # Read headers
        headers = [cell.value if cell.value else f"col_{i}" for i, cell in enumerate(ws[header_row], start=1)]
        headers_lower = [h.lower().strip() if isinstance(h, str) else h for h in headers]
        
        # Map column names
        tc_col = None
        name_col = None
        dept_col = None
        
        for i, h in enumerate(headers_lower):
            if h and ('tc' in h or 'kimlik' in h):
                tc_col = i
            elif h and ('personel' in h or 'ad' in h or 'soyad' in h):
                name_col = i
            elif h and ('departman' in h or 'bölüm' in h):
                dept_col = i
        
        if tc_col is None:
            raise ValueError("TC Kimlik No kolonu bulunamadı")
        
        # Find day columns (gun_1 to gun_31)
        day_columns = {}
        fm_columns = {}
        # Additional summary columns
        summary_columns = {}
        
        for i, h in enumerate(headers_lower):
            if not h:
                continue
            h_str = str(h).strip()
            
            # gun_1, gun_2, ... or 1, 2, 3, ...
            match = re.match(r'(?:gun[_\s]*)?(\d+)$', h_str)
            if match and 'fm' not in h_str:
                day_num = int(match.group(1))
                if 1 <= day_num <= 31:
                    day_columns[day_num] = i
            # fm_gun_1, fm_1, FM_1, ...
            elif 'fm' in h_str:
                match_fm = re.match(r'(?:fm[_\s]*(?:gun[_\s]*)?)?(\d+)$', h_str)
                if match_fm:
                    day_num = int(match_fm.group(1))
                    if 1 <= day_num <= 31:
                        fm_columns[day_num] = i
            
            # Summary/calculation columns
            if 'çalışılan' in h_str or 'calisilan' in h_str:
                summary_columns['calisilan_gun_sayisi'] = i
            elif 'ssk' in h_str and 'gun' in h_str:
                summary_columns['ssk_gun_sayisi'] = i
            elif 'yıllık' in h_str or 'yillik' in h_str:
                summary_columns['yillik_izin_gun'] = i
            elif 'izin' in h_str and 'gun' in h_str and 'yıllık' not in h_str:
                summary_columns['izin_gun_sayisi'] = i
            elif 'rapor' in h_str:
                summary_columns['rapor_gun_sayisi'] = i
            elif 'eksik' in h_str:
                summary_columns['eksik_gun_sayisi'] = i
            elif 'yarım' in h_str or 'yarim' in h_str:
                summary_columns['yarim_gun_sayisi'] = i
            elif 'toplam' in h_str and 'gun' in h_str:
                summary_columns['toplam_gun_sayisi'] = i
            elif 'normal' in h_str and 'çalışma' in h_str:
                summary_columns['normal_calismasi'] = i
            elif 'fazla' in h_str and 'çalışma' in h_str:
                summary_columns['fazla_calismasi'] = i
            elif 'gece' in h_str and 'çalışma' in h_str:
                summary_columns['gece_calismasi'] = i
            elif 'tatil' in h_str and 'çalışma' in h_str:
                summary_columns['tatil_calismasi'] = i
            elif 'sigorta' in h_str and 'girmedigi' in h_str:
                summary_columns['sigorta_girmedigi'] = i
            elif 'hafta' in h_str and 'tatil' in h_str:
                summary_columns['hafta_tatili'] = i
            elif 'resmi' in h_str and 'tatil' in h_str:
                summary_columns['resmi_tatil'] = i
            elif h_str == 'yol':
                summary_columns['yol'] = i
            elif h_str == 'prim':
                summary_columns['prim'] = i
            elif h_str == 'ikramiye':
                summary_columns['ikramiye'] = i
            elif h_str == 'bayram':
                summary_columns['bayram'] = i
            elif h_str == 'kira':
                summary_columns['kira'] = i
        
        # Process data rows
        saved_count = 0
        updated_count = 0
        skipped_count = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            row_values = [cell.value for cell in row]
            
            # Skip empty rows
            if not any(row_values):
                continue
            
            # Get TC Kimlik No
            tc_kimlik = row_values[tc_col] if tc_col < len(row_values) else None
            if not tc_kimlik:
                skipped_count += 1
                continue
            
            tc_kimlik = str(tc_kimlik).strip()
            
            # Find personnel by TC
            from app.models import Personnel
            personnel = self.db.query(Personnel).filter(
                Personnel.tc_kimlik_no == tc_kimlik
            ).first()
            
            if not personnel:
                print(f"[WARNING] Personnel not found for TC: {tc_kimlik}")
                skipped_count += 1
                continue
            
            # Find DRAFT contract (maaş hesabı için - dönemlik değil, aktif taslak sözleşme)
            from app.models import PersonnelDraftContract
            draft_contract = self.db.query(PersonnelDraftContract).filter(
                PersonnelDraftContract.personnel_id == personnel.id,
                PersonnelDraftContract.is_active == 1
            ).first()
            
            # contract_id NULL olmalı (draft contract ID'si personnel_contracts'ta yok, FK hatası verir)
            # Bordro hesaplaması draft contract'ı personnel_id üzerinden bulacak
            contract_id = None
            cost_center_id = draft_contract.cost_center_id if draft_contract else None
            
            # Check if record exists (bir personelin bir dönemde tek puantajı var)
            existing = self.repository.get_by_personnel_donem(
                self.db, personnel.id, donem
            )
            
            # Build day data
            gun_data = {}
            for day_num in range(1, 32):
                gun_key = f'gun_{day_num}'
                fm_key = f'fm_gun_{day_num}'
                
                # Status code
                if day_num in day_columns:
                    val = row_values[day_columns[day_num]] if day_columns[day_num] < len(row_values) else None
                    if val:
                        gun_data[gun_key] = str(val).strip().upper() if isinstance(val, str) else str(val)
                
                # FM value
                if day_num in fm_columns:
                    fm_val = row_values[fm_columns[day_num]] if fm_columns[day_num] < len(row_values) else None
                    if fm_val:
                        try:
                            gun_data[fm_key] = float(fm_val)
                        except:
                            pass
            
            # Add summary/calculation columns (if exist in Excel)
            for col_name, col_idx in summary_columns.items():
                if col_idx < len(row_values):
                    val = row_values[col_idx]
                    if val is not None and str(val).strip():
                        try:
                            # Decimal fields
                            if col_name in ['yarim_gun_sayisi', 'normal_calismasi', 'fazla_calismasi', 
                                           'gece_calismasi', 'tatil_calismasi', 'yol', 'prim', 
                                           'ikramiye', 'bayram', 'kira']:
                                gun_data[col_name] = float(val)
                            # Integer fields
                            else:
                                gun_data[col_name] = int(float(val))
                        except (ValueError, TypeError):
                            pass  # Skip invalid values
            
            # Calculate summary fields from gun_X values if not provided in Excel
            # Bu hesaplamalar Excel'de yoksa otomatik yapılır
            if 'calisilan_gun_sayisi' not in gun_data:
                gun_data['calisilan_gun_sayisi'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') == 'C')
            
            if 'izin_gun_sayisi' not in gun_data:
                gun_data['izin_gun_sayisi'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') == 'I')
            
            if 'rapor_gun_sayisi' not in gun_data:
                gun_data['rapor_gun_sayisi'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') == 'R')
            
            if 'eksik_gun_sayisi' not in gun_data:
                gun_data['eksik_gun_sayisi'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') == 'U')
            
            if 'resmi_tatil' not in gun_data:
                gun_data['resmi_tatil'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') == 'T')
            
            # Normal çalışma (FM toplamı)
            if 'normal_calismasi' not in gun_data:
                fm_total = sum(gun_data.get(f'fm_gun_{i}') or 0 for i in range(1, 32))
                gun_data['normal_calismasi'] = round(fm_total, 2) if fm_total > 0 else 0
            
            # Toplam gün sayısı
            if 'toplam_gun_sayisi' not in gun_data:
                gun_data['toplam_gun_sayisi'] = sum(1 for i in range(1, 32) if gun_data.get(f'gun_{i}') and gun_data.get(f'gun_{i}') != '')
            
            if existing:
                # UPDATE - contract_id de güncellenebilir (draft contract değişmiş olabilir)
                gun_data['updated_at'] = datetime.now()
                gun_data['contract_id'] = contract_id
                gun_data['cost_center_id'] = cost_center_id
                self.repository.update(self.db, existing, **gun_data)
                updated_count += 1
            else:
                # INSERT
                # Calculate days in month
                days_in_month = calendar.monthrange(year, month)[1]
                
                new_record = PersonnelPuantajGrid(
                    personnel_id=personnel.id,
                    cost_center_id=cost_center_id,
                    donem=donem,
                    yil=year,
                    ay=month,
                    ayin_toplam_gun_sayisi=days_in_month,
                    **gun_data
                )
                self.repository.create(self.db, new_record)
                saved_count += 1
        
        self.db.commit()
        
        return {
            "success": True,
            "donem": donem,
            "saved": saved_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "total": saved_count + updated_count
        }
    
    def create_template_excel(self, donem: str, cost_center_id: Optional[int] = None) -> bytes:
        """
        Create Excel template for puantaj grid
        
        Args:
            donem: Period (YYYY-MM)
            cost_center_id: Optional cost center ID to filter personnel
            
        Returns:
            Excel file bytes
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        year, month = map(int, donem.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Puantaj"
        
        # Status codes legend sheet
        legend_ws = wb.create_sheet("Durum Kodları")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Build headers
        headers = ["Personel", "TC Kimlik No", "Departman", "Net Maaş", "FM Oranı", "Tatil Oranı"]
        
        # Add day columns - YAN YANA: Gün ve FM
        for i in range(1, days_in_month + 1):
            headers.append(f"{i}")      # Gün kolonu
            headers.append(f"FM_{i}")   # FM kolonu hemen yanında
        
        # Toplam Kazanç kolonu
        headers.append("Toplam Kazanç")
        
        # Write headers
        col_idx = 1
        for header in headers:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            col_idx += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Personel
        ws.column_dimensions['B'].width = 15  # TC
        ws.column_dimensions['C'].width = 20  # Departman
        ws.column_dimensions['D'].width = 12  # Net Maaş
        ws.column_dimensions['E'].width = 10  # FM Oranı
        ws.column_dimensions['F'].width = 10  # Tatil Oranı
        
        # Day ve FM columns - yan yana (7'den başlar)
        for i in range(days_in_month):
            day_col_idx = 7 + (i * 2)      # Gün kolonu
            fm_col_idx = 7 + (i * 2) + 1   # FM kolonu
            ws.column_dimensions[get_column_letter(day_col_idx)].width = 4   # Gün dar
            ws.column_dimensions[get_column_letter(fm_col_idx)].width = 5    # FM biraz geniş
        
        # Toplam Kazanç kolonu
        kazanc_col_idx = 7 + (days_in_month * 2)
        ws.column_dimensions[get_column_letter(kazanc_col_idx)].width = 15
        
        # Get personnel data - tabloda gösterilen personeller
        grid_data = self.get_grid_data(donem, cost_center_id)
        personnel_list = grid_data.get('records', [])
        
        # Filter out header rows
        personnel_list = [p for p in personnel_list if p.get('row_type') not in ['header', 'taseron_header']]
        
        # Gri arka plan sigortası olmayan günler için (daha belirgin gri)
        disabled_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Write personnel data
        current_row = 2
        for person in personnel_list:
            # Personnel info - DOLU GELSİN
            ws.cell(row=current_row, column=1, value=person.get('adi_soyadi', ''))
            ws.cell(row=current_row, column=2, value=person.get('tc_kimlik_no', ''))
            ws.cell(row=current_row, column=3, value=person.get('departman', ''))
            
            # Ücret nevi - backend'de biliyoruz, kullanıcıya göstermiyoruz
            ucret_nevi = person.get('ucret_nevi', None)
            
            # Net Maaş - Draft contract'tan gelen net_ucret
            net_maas = person.get('maas2_tutar', None)
            net_maas_cell = ws.cell(row=current_row, column=4, value=float(net_maas) if net_maas else '')
            if net_maas:
                net_maas_cell.number_format = '#,##0.00'  # Para formatı
            net_maas_cell.alignment = Alignment(horizontal="right")
            net_maas_cell.border = thin_border
            
            # FM Oranı - Draft contract'tan
            fm_orani = person.get('fm_orani', None)
            fm_orani_cell = ws.cell(row=current_row, column=5, value=float(fm_orani) if fm_orani else '')
            if fm_orani:
                fm_orani_cell.number_format = '0.0'  # Ondalık formatı
            fm_orani_cell.alignment = Alignment(horizontal="center")
            fm_orani_cell.border = thin_border
            
            # Tatil Oranı - Draft contract'tan
            tatil_orani = person.get('tatil_orani', None)
            tatil_orani_cell = ws.cell(row=current_row, column=6, value=float(tatil_orani) if tatil_orani else '')
            if tatil_orani:
                tatil_orani_cell.number_format = '0.0'  # Ondalık formatı
            tatil_orani_cell.alignment = Alignment(horizontal="center")
            tatil_orani_cell.border = thin_border
            
            # Ücret nevi kontrolü - sabit aylık olanlar için puantajı otomatik doldur
            ucret_nevi = person.get('ucret_nevi', None)
            auto_fill_puantaj = (ucret_nevi == 'sabit aylik')
            
            # Pazar günleri için açık sarı renk
            sunday_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            # Day and FM values - YAN YANA
            for day_num in range(1, days_in_month + 1):
                # Günün tarihini hesapla
                day_date = date(year, month, day_num)
                day_of_week = day_date.weekday()  # 0=Pazartesi, 6=Pazar
                is_sunday = (day_of_week == 6)
                
                # Bu günün değerini al
                gun_key = f'gun_{day_num}'
                day_value_from_data = person.get(gun_key, None)
                
                # Bu gün sigortası yok mu? (değeri '-' ise)
                is_not_insured = (day_value_from_data == '-')
                
                # Kolon indeksleri (7'den başlıyor: G, H, I, J, ...)
                day_col_idx = 7 + ((day_num - 1) * 2)      # Gün kolonu
                fm_col_idx = 7 + ((day_num - 1) * 2) + 1   # FM kolonu
                
                # Gün değeri belirleme
                if is_not_insured:
                    day_value = '-'  # Sigorta girmediği günler
                elif auto_fill_puantaj:
                    # Sabit aylık personeller için otomatik değer
                    if is_sunday:
                        day_value = 'H'  # Hafta tatili
                    else:
                        day_value = 'N'  # Normal çalışma
                else:
                    day_value = ''  # Boş bırak (manuel girilecek)
                
                # Gün kolonu
                day_cell = ws.cell(row=current_row, column=day_col_idx, value=day_value)
                day_cell.alignment = Alignment(horizontal="center")
                day_cell.border = thin_border
                
                if is_not_insured:
                    # Sigortası olmayan günler: gri arka plan + '-' karakteri
                    day_cell.fill = disabled_fill
                elif is_sunday:
                    # Pazar: sarı arka plan
                    day_cell.fill = sunday_fill
                
                # FM kolonu
                fm_value = ''  # Boş bırak (sigortası olmayan günlerde de boş)
                fm_cell = ws.cell(row=current_row, column=fm_col_idx, value=fm_value)
                fm_cell.alignment = Alignment(horizontal="center")
                fm_cell.border = thin_border
                
                if is_not_insured:
                    # Sigortası olmayan günler: gri arka plan
                    fm_cell.fill = disabled_fill
                elif is_sunday:
                    # Pazar: sarı arka plan
                    fm_cell.fill = sunday_fill
            
            # Toplam Kazanç formülü - en sağa
            kazanc_col_idx = 7 + (days_in_month * 2)
            
            # Gün kolonları range (G'den başlar, her 2 kolonda bir - sadece gün kolonları)
            first_day_col = get_column_letter(7)
            last_day_col = get_column_letter(7 + ((days_in_month - 1) * 2))
            day_range = f"{first_day_col}{current_row}:{last_day_col}{current_row}"
            
            # FM kolonları - Her 2 kolonda bir (H, J, L, N, P...)
            # DOĞRU YÖNTEM: Her FM kolonunu ayrı ayrı topla
            fm_cells_list = [f"{get_column_letter(8 + i*2)}{current_row}" for i in range(days_in_month)]
            fm_toplam_formula = "+".join(fm_cells_list)
            
            # Sayımlar
            calisilan_gun_say = f'COUNTIF({day_range},"N")'
            yarim_gun_say = f'COUNTIF({day_range},"Y")'
            hafta_tatil_say = f'COUNTIF({day_range},"H")'
            resmi_tatil_say = f'COUNTIF({day_range},"T")'
            tatil_calisma_say = f'COUNTIF({day_range},"M")'
            ucretli_izin_say = f'COUNTIF({day_range},"İ")'
            yillik_izin_say = f'COUNTIF({day_range},"S")'
            rapor_say = f'COUNTIF({day_range},"R")'
            eksik_gun_say = f'COUNTIF({day_range},"E")'
            sigorta_girmedigi_say = f'COUNTIF({day_range},"-")'
            
            # FM/EM hesabı - Pozitif ise FM, Negatif ise EM
            fm_sum_base = f'SUM({fm_toplam_formula})'
            fm_toplam = f'IF({fm_sum_base}>0,{fm_sum_base},0)'  # Pozitif kısım
            em_toplam = f'IF({fm_sum_base}<0,ABS({fm_sum_base}),0)'  # Negatif kısım (mutlak değer)
            
            # FM ve Tatil oranları (boşsa default değerler)
            fm_orani_formula = f"IF(ISBLANK(E{current_row}),1.5,E{current_row})"
            tatil_orani_formula = f"IF(ISBLANK(F{current_row}),1,F{current_row})"
            
            # FORMÜL
            if ucret_nevi == 'gunluk':
                # GÜNLÜKÇÜ: D kolonu = günlük ücret
                gunluk_ucret = f'D{current_row}'
                formula = f"=({gunluk_ucret})*(({calisilan_gun_say}+{yarim_gun_say}*0.5)+({fm_toplam}/8)*({fm_orani_formula})-({em_toplam}/8)+({hafta_tatil_say}+{resmi_tatil_say}+{tatil_calisma_say})+{tatil_calisma_say}*({tatil_orani_formula})+({yillik_izin_say}))"
            else:
                # AYLIKÇI/SABİT AYLIK: D kolonu = aylık net maaş, gunluk = D/30
                gunluk_ucret = f'D{current_row}/30'
                
                # Tatiller toplamı
                tatiller_toplam = f'({hafta_tatil_say}+{resmi_tatil_say}+{tatil_calisma_say})'
                
                # ÖZEL DURUM (Tam Ay): Ay≠30 VE Eksik=0 VE Sigortasız=0 VE Rapor=0 VE Yarım=0
                # normal_calismasi = 30 - tatiller - izin_sinirli - yillik_izin
                # Toplam = (30 - tatiller - izin - yillik_izin)*gunluk + izin*gunluk + FM - EM + tatiller*gunluk + M*tatil_orani*gunluk + yillik_izin*gunluk
                # Basitleştirilmiş: 30*gunluk + FM - EM + M*tatil_orani*gunluk
                
                # NORMAL: normal_calismasi = N + Y*0.5
                # Toplam = (N+Y*0.5)*gunluk + FM - EM + tatiller*gunluk + M*tatil_orani*gunluk + yillik_izin*gunluk
                
                formula = f"=IF(AND({days_in_month}<>30,{eksik_gun_say}=0,{sigorta_girmedigi_say}=0,{rapor_say}=0,{yarim_gun_say}=0),({gunluk_ucret})*(30+{tatil_calisma_say}*{tatil_orani_formula}+{fm_toplam}/8*{fm_orani_formula}-{em_toplam}/8),({gunluk_ucret})*(({calisilan_gun_say}+{yarim_gun_say}*0.5)+({hafta_tatil_say}+{resmi_tatil_say}+{tatil_calisma_say})+{tatil_calisma_say}*{tatil_orani_formula}+({yillik_izin_say})+{fm_toplam}/8*{fm_orani_formula}-{em_toplam}/8))"
            
            kazanc_cell = ws.cell(row=current_row, column=kazanc_col_idx, value=formula)
            kazanc_cell.number_format = '#,##0.00'  # Para formatı
            kazanc_cell.alignment = Alignment(horizontal="right")
            kazanc_cell.border = thin_border
            
            current_row += 1
        
        # Freeze first row and first 6 columns (Personel, TC, Departman, Net Maaş, FM Oranı, Tatil Oranı)
        ws.freeze_panes = "G2"
        
        # =============================================
        # KAZANÇLAR SAYFASI
        # =============================================
        kazanc_ws = wb.create_sheet("Kazançlar", 1)  # Puantaj'dan sonra, Durum Kodları'ndan önce
        
        # Kazançlar sayfası headers
        kazanc_headers = [
            "Personel", "Ücret Nevi", "Net Ücret", "Günlük Ücret",
            "Normal Gün", "Normal Kazanç", "İzin Gün", "İzin Kazanç",
            "Mesai Saat", "Mesai Kazanç", "Eksik Mesai Saat", "Eksik Mesai Kesinti",
            "Tatil Gün", "Tatil Kazanç", "Tatil Mesai Gün", "Tatil Mesai Kazanç",
            "Yıllık İzin Gün", "Yıllık İzin Kazanç",
            "Yol", "Prim", "İkramiye", "Bayram", "Kira", "Toplam Kazanç",
            "Çalışılan Gün", "Yarım Gün", "Eksik Gün", "Sigortasız", "Rapor", "Ayın Toplam Günü"
        ]
        
        # Write headers
        for col_idx, header in enumerate(kazanc_headers, start=1):
            cell = kazanc_ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        kazanc_widths = [25, 12, 12, 12, 10, 12, 10, 12, 10, 12, 14, 14, 10, 12, 14, 16, 14, 16, 12, 12, 12, 12, 12, 15, 12, 10, 10, 10, 10, 16]
        for i, width in enumerate(kazanc_widths, 1):
            kazanc_ws.column_dimensions[get_column_letter(i)].width = width
        
        # Write personnel data to Kazançlar sheet
        kazanc_row = 2
        for person in personnel_list:
            # Gün kolonları aralığını hesapla (dinamik - ayın gün sayısına göre)
            # İlk gün kolonu: G (7)
            # Son gün kolonu: 7 + (days_in_month - 1) * 2
            first_day_col = 'G'
            last_day_col_idx = 7 + (days_in_month - 1) * 2
            last_day_col = get_column_letter(last_day_col_idx)
            day_range = f'Puantaj!{first_day_col}{kazanc_row}:{last_day_col}{kazanc_row}'
            
            # A: Personel adı (Puantaj'dan referans)
            kazanc_ws.cell(row=kazanc_row, column=1, value=f"='Puantaj'!A{kazanc_row}").border = thin_border
            
            # B: Ücret Nevi
            ucret_nevi_cell = kazanc_ws.cell(row=kazanc_row, column=2, value=person.get('ucret_nevi', ''))
            ucret_nevi_cell.alignment = Alignment(horizontal="center")
            ucret_nevi_cell.border = thin_border
            
            # C: Net Ücret (Puantaj'dan referans)
            kazanc_ws.cell(row=kazanc_row, column=3, value=f"='Puantaj'!D{kazanc_row}").number_format = '#,##0.00'
            kazanc_ws.cell(row=kazanc_row, column=3).border = thin_border
            
            # D: Günlük Ücret
            gunluk_ucret_cell = kazanc_ws.cell(row=kazanc_row, column=4, value=f'=IF(B{kazanc_row}="gunluk",C{kazanc_row},C{kazanc_row}/30)')
            gunluk_ucret_cell.number_format = '#,##0.00'
            gunluk_ucret_cell.border = thin_border
            
            # Y-AD: Yardımcı kolonlar (25-30)
            # Y: Çalışılan Gün (N kodları)
            kazanc_ws.cell(row=kazanc_row, column=25, value=f'=COUNTIF({day_range},"N")').border = thin_border
            
            # Z: Yarım Gün (Y kodları)
            kazanc_ws.cell(row=kazanc_row, column=26, value=f'=COUNTIF({day_range},"Y")').border = thin_border
            
            # AA: Eksik Gün (E kodları)
            kazanc_ws.cell(row=kazanc_row, column=27, value=f'=COUNTIF({day_range},"E")').border = thin_border
            
            # AB: Sigortasız (- kodları)
            kazanc_ws.cell(row=kazanc_row, column=28, value=f'=COUNTIF({day_range},"-")').border = thin_border
            
            # AC: Rapor (R kodları)
            kazanc_ws.cell(row=kazanc_row, column=29, value=f'=COUNTIF({day_range},"R")').border = thin_border
            
            # AD: Ayın Toplam Günü (backend'den)
            kazanc_ws.cell(row=kazanc_row, column=30, value=days_in_month).border = thin_border
            
            # E: Normal Gün (tam ay formülü ile)
            normal_gun_formula = f'=IF(AND(OR(B{kazanc_row}="aylik",B{kazanc_row}="sabit aylik"),AA{kazanc_row}=0,AD{kazanc_row}<>30,AB{kazanc_row}=0,AC{kazanc_row}=0,Z{kazanc_row}=0),30-M{kazanc_row}-G{kazanc_row}-Q{kazanc_row},Y{kazanc_row}+Z{kazanc_row}*0.5)'
            kazanc_ws.cell(row=kazanc_row, column=5, value=normal_gun_formula).border = thin_border
            
            # F: Normal Kazanç
            normal_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=6, value=f'=E{kazanc_row}*D{kazanc_row}')
            normal_kazanc_cell.number_format = '#,##0.00'
            normal_kazanc_cell.border = thin_border
            
            # G: İzin Gün (max 30)
            kazanc_ws.cell(row=kazanc_row, column=7, value=f'=MIN(COUNTIF({day_range},"İ"),30)').border = thin_border
            
            # H: İzin Kazanç
            izin_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=8, value=f'=G{kazanc_row}*D{kazanc_row}')
            izin_kazanc_cell.number_format = '#,##0.00'
            izin_kazanc_cell.border = thin_border
            
            # I: Mesai Saat (FM toplamı, pozitif kısım)
            # FM kolonlarını tek tek topla (H, J, L, N, ... 31 gün için)
            fm_cols = [get_column_letter(8 + i*2) for i in range(days_in_month)]
            fm_sum_formula = '+'.join([f'Puantaj!{col}{kazanc_row}' for col in fm_cols])
            mesai_saat_cell = kazanc_ws.cell(row=kazanc_row, column=9, value=f'=MAX(0,SUM({fm_sum_formula}))')
            mesai_saat_cell.number_format = '0.0'
            mesai_saat_cell.border = thin_border
            
            # J: Mesai Kazanç
            mesai_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=10, value=f'=(I{kazanc_row}/8)*D{kazanc_row}*Puantaj!E{kazanc_row}')
            mesai_kazanc_cell.number_format = '#,##0.00'
            mesai_kazanc_cell.border = thin_border
            
            # K: Eksik Mesai Saat (FM toplamı, negatif kısım)
            eksik_mesai_saat_cell = kazanc_ws.cell(row=kazanc_row, column=11, value=f'=MAX(0,-SUM({fm_sum_formula}))')
            eksik_mesai_saat_cell.number_format = '0.0'
            eksik_mesai_saat_cell.border = thin_border
            
            # L: Eksik Mesai Kesinti
            eksik_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=12, value=f'=(K{kazanc_row}/8)*D{kazanc_row}')
            eksik_kazanc_cell.number_format = '#,##0.00'
            eksik_kazanc_cell.border = thin_border
            
            # M: Tatil Gün (H + T + M)
            kazanc_ws.cell(row=kazanc_row, column=13, value=f'=COUNTIF({day_range},"H")+COUNTIF({day_range},"T")+COUNTIF({day_range},"M")').border = thin_border
            
            # N: Tatil Kazanç
            tatil_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=14, value=f'=M{kazanc_row}*D{kazanc_row}')
            tatil_kazanc_cell.number_format = '#,##0.00'
            tatil_kazanc_cell.border = thin_border
            
            # O: Tatil Mesai Gün (M kodları)
            kazanc_ws.cell(row=kazanc_row, column=15, value=f'=COUNTIF({day_range},"M")').border = thin_border
            
            # P: Tatil Mesai Kazanç
            tatil_mesai_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=16, value=f'=O{kazanc_row}*D{kazanc_row}*Puantaj!F{kazanc_row}')
            tatil_mesai_kazanc_cell.number_format = '#,##0.00'
            tatil_mesai_kazanc_cell.border = thin_border
            
            # Q: Yıllık İzin Gün (S kodları)
            kazanc_ws.cell(row=kazanc_row, column=17, value=f'=COUNTIF({day_range},"S")').border = thin_border
            
            # R: Yıllık İzin Kazanç
            yillik_izin_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=18, value=f'=Q{kazanc_row}*D{kazanc_row}')
            yillik_izin_kazanc_cell.number_format = '#,##0.00'
            yillik_izin_kazanc_cell.border = thin_border
            
            # S-W: Ek Ödemeler (Backend'den)
            yol_cell = kazanc_ws.cell(row=kazanc_row, column=19, value=float(person.get('yol', 0)))
            yol_cell.number_format = '#,##0.00'
            yol_cell.border = thin_border
            
            prim_cell = kazanc_ws.cell(row=kazanc_row, column=20, value=float(person.get('prim', 0)))
            prim_cell.number_format = '#,##0.00'
            prim_cell.border = thin_border
            
            ikramiye_cell = kazanc_ws.cell(row=kazanc_row, column=21, value=float(person.get('ikramiye', 0)))
            ikramiye_cell.number_format = '#,##0.00'
            ikramiye_cell.border = thin_border
            
            bayram_cell = kazanc_ws.cell(row=kazanc_row, column=22, value=float(person.get('bayram', 0)))
            bayram_cell.number_format = '#,##0.00'
            bayram_cell.border = thin_border
            
            kira_cell = kazanc_ws.cell(row=kazanc_row, column=23, value=float(person.get('kira', 0)))
            kira_cell.number_format = '#,##0.00'
            kira_cell.border = thin_border
            
            # X: Toplam Kazanç
            toplam_kazanc_cell = kazanc_ws.cell(row=kazanc_row, column=24, value=f'=F{kazanc_row}+H{kazanc_row}+J{kazanc_row}-L{kazanc_row}+N{kazanc_row}+P{kazanc_row}+R{kazanc_row}+S{kazanc_row}+T{kazanc_row}+U{kazanc_row}+V{kazanc_row}+W{kazanc_row}')
            toplam_kazanc_cell.number_format = '#,##0.00'
            toplam_kazanc_cell.font = Font(bold=True)
            toplam_kazanc_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            toplam_kazanc_cell.border = thin_border
            
            kazanc_row += 1
        
        # Freeze first row and first 4 columns
        kazanc_ws.freeze_panes = "E2"
        
        # Legend sheet
        legend_ws.column_dimensions['A'].width = 8
        legend_ws.column_dimensions['B'].width = 40
        legend_ws.column_dimensions['C'].width = 50
        
        legend_ws.cell(row=1, column=1, value="KOD").font = Font(bold=True, size=12)
        legend_ws.cell(row=1, column=2, value="AÇIKLAMA").font = Font(bold=True, size=12)
        legend_ws.cell(row=1, column=3, value="NOTLAR").font = Font(bold=True, size=12)
        
        status_codes = [
            ("N", "Normal Çalışma", "Günlük normal çalışma günü"),
            ("H", "Hafta Tatili", "Cumartesi/Pazar gibi hafta tatili günleri"),
            ("T", "Resmi Tatil", "Resmi tatil günleri (bayram, milli tatil vb.)"),
            ("İ", "İzinli", "Mazeret izni"),
            ("S", "Yıllık İzin", "Yıllık izin kullanımı"),
            ("R", "Raporlu", "Sağlık raporu olan günler"),
            ("E", "Eksik Gün", "SSK'ya girilmeyen eksik gün"),
            ("Y", "Yarım Gün", "Yarım gün çalışma (0.5 gün)"),
            ("M", "Tatil Çalışması", "Hafta tatili/resmi tatilde çalışma"),
            ("", "", ""),
            ("FM_X", "Fazla Mesai", "Gün kolonunun yanındaki FM kolonuna saat cinsinden girilir (örn: 2, 4, 8)"),
        ]
        
        for idx, (code, desc, note) in enumerate(status_codes, start=2):
            legend_ws.cell(row=idx, column=1, value=code).font = Font(bold=True if code else False)
            legend_ws.cell(row=idx, column=2, value=desc)
            legend_ws.cell(row=idx, column=3, value=note).font = Font(italic=True, color="666666")
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
