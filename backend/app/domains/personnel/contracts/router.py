"""
Contracts Router (V2)
FastAPI endpoints for personnel contracts - Full V2 migration
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import and_, or_, desc, asc, func

from app.core.database import get_db
from app.domains.auth.dependencies import get_current_user
from app.schemas.auth import UserInDB
from app.domains.personnel.contracts.schemas import (
    PersonnelContractCreate,
    PersonnelContractUpdate,
    PersonnelContractResponse,
    PersonnelContractList
)
from app.models import PersonnelContract
from app.models import Personnel
from app.models import CostCenter
from app.models import Contact
from app.models import MonthlyPersonnelRecord

router = APIRouter(tags=["Personnel - Contracts (V2)"])


def safe_enum_value(enum_val, default=''):
    """Enum veya string değeri güvenli şekilde string'e çevir"""
    if enum_val is None:
        return default
    if isinstance(enum_val, str):
        return enum_val
    return enum_val.value if hasattr(enum_val, 'value') else str(enum_val)


@router.get("/periods")
def get_periods(
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Dönemleri listele (monthly_personnel_records'dan)
    """
    periods = db.query(
        MonthlyPersonnelRecord.donem,
        MonthlyPersonnelRecord.yil,
        MonthlyPersonnelRecord.ay,
        func.max(MonthlyPersonnelRecord.id).label('latest_id')
    ).group_by(
        MonthlyPersonnelRecord.donem,
        MonthlyPersonnelRecord.yil,
        MonthlyPersonnelRecord.ay
    ).order_by(
        MonthlyPersonnelRecord.yil.desc(),
        MonthlyPersonnelRecord.ay.desc()
    ).all()
    
    period_list = [
        {
            "donem": p.donem,
            "yil": p.yil,
            "ay": p.ay,
            "label": f"{p.donem} ({p.yil}/{p.ay:02d})",
            "is_latest": False
        }
        for p in periods
    ]
    
    # En son dönemi işaretle
    if period_list:
        period_list[0]['is_latest'] = True
    
    return {
        "periods": period_list,
        "latest_period": period_list[0] if period_list else None
    }


@router.get("/list", response_model=PersonnelContractList)
def list_contracts(
    personnel_id: Optional[int] = None,
    is_active: Optional[int] = None,
    donem: Optional[str] = None,
    cost_center_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 100,
    order_by: Optional[str] = None,
    order_direction: Optional[str] = 'desc',
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Sözleşmeleri listele - dönem ve cost center filtrelemesi ile sıralama
    """
    # Personnel, Contact, CostCenter ve MonthlyPersonnelRecord tablolarıyla join yap
    query = db.query(
        PersonnelContract,
        Personnel.ad.label('personnel_ad'),
        Personnel.soyad.label('personnel_soyad'),
        Contact.name.label('taseron_name'),
        CostCenter.name.label('cost_center_name'),
        MonthlyPersonnelRecord.meslek_adi.label('meslek_adi')
    ).join(
        Personnel, 
        PersonnelContract.personnel_id == Personnel.id
    ).outerjoin(
        Contact,
        PersonnelContract.taseron_id == Contact.id
    ).outerjoin(
        CostCenter,
        PersonnelContract.cost_center_id == CostCenter.id
    ).outerjoin(
        MonthlyPersonnelRecord,
        PersonnelContract.monthly_personnel_records_id == MonthlyPersonnelRecord.id
    )
    
    if personnel_id:
        query = query.filter(PersonnelContract.personnel_id == personnel_id)
    
    if is_active is not None:
        query = query.filter(PersonnelContract.is_active == is_active)
    
    if cost_center_id:
        query = query.filter(PersonnelContract.cost_center_id == cost_center_id)
    
    # Dönem filtresi - seçili dönemde aktif olan sözleşmeler
    if donem:
        from datetime import datetime
        yil, ay = map(int, donem.split('-'))
        # Dönemin son günü
        if ay == 12:
            next_month = datetime(yil + 1, 1, 1)
        else:
            next_month = datetime(yil, ay + 1, 1)
        donem_son_gun = (next_month - pd.Timedelta(days=1)).date()
        donem_ilk_gun = datetime(yil, ay, 1).date()
        
        query = query.filter(
            PersonnelContract.ise_giris_tarihi <= donem_son_gun,
            or_(
                PersonnelContract.isten_cikis_tarihi == None,
                PersonnelContract.isten_cikis_tarihi >= donem_ilk_gun
            )
        )
    
    # Toplam sayı
    total = query.count()
    
    # Sıralama - personnel_ad ve taseron_name için ilgili tablolardan sırala
    if order_by:
        if order_by == 'personnel_ad':
            # Ad ve soyad'a göre sırala
            if order_direction == 'asc':
                query = query.order_by(asc(Personnel.ad), asc(Personnel.soyad))
            else:
                query = query.order_by(desc(Personnel.ad), desc(Personnel.soyad))
        elif order_by == 'taseron_name':
            # Taşeron adına göre sırala (NULL değerler sona)
            if order_direction == 'asc':
                query = query.order_by(asc(Contact.name))
            else:
                query = query.order_by(desc(Contact.name))
        else:
            order_column = getattr(PersonnelContract, order_by, PersonnelContract.ise_giris_tarihi)
            if order_direction == 'asc':
                query = query.order_by(asc(order_column))
            else:
                query = query.order_by(desc(order_column))
    else:
        query = query.order_by(desc(PersonnelContract.ise_giris_tarihi))
    
    # Pagination
    skip = (page - 1) * page_size
    results = query.offset(skip).limit(page_size).all()
    
    # Response model için dönüştür - Sadece resmi sözleşme bilgileri
    items = []
    for contract, personnel_ad, personnel_soyad, taseron_name, cost_center_name, meslek_adi in results:
        contract_dict = {
            'id': contract.id,
            'personnel_id': contract.personnel_id,
            'tc_kimlik_no': contract.tc_kimlik_no,
            'bolum': safe_enum_value(contract.bolum),
            'ise_giris_tarihi': contract.ise_giris_tarihi,
            'isten_cikis_tarihi': contract.isten_cikis_tarihi,
            'kanun_tipi': safe_enum_value(contract.kanun_tipi),
            'net_brut': safe_enum_value(contract.net_brut),
            'ucret': contract.ucret,
            'iban': contract.iban,
            'taseron': contract.taseron,
            'taseron_id': contract.taseron_id,
            'departman': safe_enum_value(contract.departman),
            'cost_center_id': contract.cost_center_id,
            'monthly_personnel_records_id': contract.monthly_personnel_records_id,
            'is_active': contract.is_active,
            'created_at': contract.created_at,
            'updated_at': contract.updated_at,
            'personnel_ad': personnel_ad,
            'personnel_soyad': personnel_soyad,
            'taseron_name': taseron_name,
            'cost_center_name': cost_center_name,
            'meslek_adi': meslek_adi
        }
        
        items.append(PersonnelContractResponse(**contract_dict))
    
    return PersonnelContractList(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


@router.get("/{contract_id}", response_model=PersonnelContractResponse)
def get_contract(
    contract_id: int,
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Sözleşme detayını getir
    """
    contract = db.query(PersonnelContract).filter(PersonnelContract.id == contract_id).first()
    
    if not contract:
        raise HTTPException(status_code=404, detail="Sözleşme bulunamadı")
    
    return contract


@router.post("/create", response_model=PersonnelContractResponse)
def create_contract(
    data: PersonnelContractCreate,
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Yeni sözleşme oluştur
    NOT: ucret_nevi, maas2_tutar, fm_orani, tatil_orani, calisma_takvimi, maas_hesabi
    artık personnel_draft_contracts tablosunda - Draft Contracts API kullanılmalı
    """
    # Personnel var mı kontrol et
    personnel = db.query(Personnel).filter(Personnel.id == data.personnel_id).first()
    if not personnel:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    
    # Otomatik alanları personnel tablosundan doldur
    contract_data = data.model_dump()
    contract_data['tc_kimlik_no'] = personnel.tc_kimlik_no
    contract_data['iban'] = personnel.iban
    
    # Sözleşme oluştur
    contract = PersonnelContract(**contract_data)
    db.add(contract)
    db.commit()
    db.refresh(contract)
    
    return contract


@router.put("/{contract_id}", response_model=PersonnelContractResponse)
def update_contract(
    contract_id: int,
    data: PersonnelContractUpdate,
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Sözleşme güncelle
    NOT: ucret_nevi, maas2_tutar, fm_orani, tatil_orani, calisma_takvimi, maas_hesabi
    artık personnel_draft_contracts tablosunda - Draft Contracts API kullanılmalı
    """
    try:
        contract = db.query(PersonnelContract).filter(PersonnelContract.id == contract_id).first()
        
        if not contract:
            raise HTTPException(status_code=404, detail="Sözleşme bulunamadı")
        
        # Güncellemeleri uygula
        update_data = data.model_dump(exclude_unset=True)
        
        # Read-only alanları çıkar (bunlar PersonnelContract tablosunda yok)
        read_only_fields = ['personnel_ad', 'personnel_soyad', 'cost_center_name', 'taseron_name', 'created_at', 'updated_at', 'created_by', 'updated_by']
        for field in read_only_fields:
            update_data.pop(field, None)
        
        # Boş string'leri NULL'a çevir (ENUM alanları için)
        # NOT: ucret_nevi, calisma_takvimi, maas_hesabi artık draft_contracts'ta - şema üzerinden gelmeyecek
        enum_fields = ['departman', 'kanun_tipi']
        for field in enum_fields:
            if field in update_data and update_data[field] == '':
                update_data[field] = None
        
        for key, value in update_data.items():
            if value is not None and hasattr(contract, key):
                setattr(contract, key, value)
            elif value is None and hasattr(contract, key):
                # NULL değerleri de set et
                setattr(contract, key, None)
        
        # NOT: ucret_nevi, maas2_tutar, fm_orani, tatil_orani, calisma_takvimi, maas_hesabi
        # artık personnel_draft_contracts tablosunda - Draft Contracts API kullanılmalı
        
        db.commit()
        db.refresh(contract)
        
        return contract
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database operation failed: {str(e)}")


@router.get("/template")
async def download_contracts_template(
    donem: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Personel sözleşme şablonu indir
    Seçili dönemdeki sicil kayıtlarına göre mevcut sözleşme bilgilerini döndürür
    
    Args:
        donem: Dönem (YYYY-MM formatında, örn: 2025-11). Verilmezse tüm aktif sözleşmeler
    """
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel Sözleşmeleri"
    
    # Header
    headers = [
        'TC Kimlik No', 'Ad', 'Soyad', 'Bölüm', 'İşe Giriş Tarihi', 'İşten Çıkış Tarihi',
        'Ücret Nevi', 'Kanun Tipi', 'Çalışma Takvimi', 'Maaş1 Tip', 'Maaş1 Tutar', 'Maaş2 Tutar',
        'Maaş Hesabı', 'IBAN', 'FM Oranı', 'Tatil Oranı', 'Taşeron', 'Taşeron ID',
        'Departman', 'Masraf Merkezi Kodu'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Veri kaynağı: Seçili dönem varsa o dönemdeki sicil kayıtları, yoksa tüm aktif sözleşmeler
    row = 2
    
    if donem:
        # Seçili dönemdeki sicil kayıtlarını al
        sicil_records = db.query(MonthlyPersonnelRecord).filter(
            MonthlyPersonnelRecord.donem == donem
        ).all()
        
        for sicil in sicil_records:
            # Bu sicil kaydına bağlı sözleşmeyi bul (draft_contract ile)
            from sqlalchemy.orm import joinedload
            contract = db.query(PersonnelContract).options(
                joinedload(PersonnelContract.draft_contract)
            ).filter(
                PersonnelContract.id == sicil.contract_id
            ).first() if sicil.contract_id else None
            
            # Personel bilgilerini al
            personnel = db.query(Personnel).filter(
                Personnel.id == sicil.personnel_id
            ).first()
            
            if not personnel:
                continue
            
            # Masraf merkezi bilgisi
            cost_center = None
            if contract and contract.cost_center_id:
                cost_center = db.query(CostCenter).filter(
                    CostCenter.id == contract.cost_center_id
                ).first()
            
            ws.cell(row=row, column=1, value=str(personnel.tc_kimlik_no) if personnel.tc_kimlik_no else '')
            ws.cell(row=row, column=2, value=str(personnel.ad) if personnel.ad else '')
            ws.cell(row=row, column=3, value=str(personnel.soyad) if personnel.soyad else '')
            ws.cell(row=row, column=4, value=str(sicil.bolum) if sicil.bolum else '')
            ws.cell(row=row, column=5, value=sicil.ise_giris_tarihi.strftime('%Y-%m-%d') if sicil.ise_giris_tarihi else '')
            ws.cell(row=row, column=6, value=sicil.isten_cikis_tarihi.strftime('%Y-%m-%d') if sicil.isten_cikis_tarihi else '')
            # Draft contract verisi al (varsa)
            draft = contract.draft_contract if (contract and hasattr(contract, 'draft_contract')) else None
            
            ws.cell(row=row, column=7, value=safe_enum_value(draft.ucret_nevi if draft else None, 'aylik'))
            ws.cell(row=row, column=8, value=safe_enum_value(contract.kanun_tipi if contract else None, '4857'))
            ws.cell(row=row, column=9, value=safe_enum_value(draft.calisma_takvimi if draft else None, 'atipi'))
            ws.cell(row=row, column=10, value=safe_enum_value(contract.net_brut if contract else None, sicil.net_brut if sicil.net_brut else 'B'))
            ws.cell(row=row, column=11, value=float(contract.ucret) if (contract and contract.ucret) else (float(sicil.ucret) if sicil.ucret else 0))
            ws.cell(row=row, column=12, value=float(draft.net_ucret) if (draft and draft.net_ucret) else 0)
            ws.cell(row=row, column=13, value=safe_enum_value(draft.maas_hesabi if draft else None, 'tipa'))
            ws.cell(row=row, column=14, value=str(personnel.iban) if personnel.iban else '')
            ws.cell(row=row, column=15, value=float(draft.fm_orani) if (draft and draft.fm_orani) else 1.0)
            ws.cell(row=row, column=16, value=float(draft.tatil_orani) if (draft and draft.tatil_orani) else 1.0)
            ws.cell(row=row, column=17, value='Evet' if (contract and contract.taseron) else 'Hayır')
            ws.cell(row=row, column=18, value=int(contract.taseron_id) if (contract and contract.taseron_id) else '')
            ws.cell(row=row, column=19, value=safe_enum_value(contract.departman if contract else None, ''))
            ws.cell(row=row, column=20, value=str(cost_center.code) if cost_center else '')
            row += 1
    else:
        # Tüm aktif sözleşmeler (draft_contract ile)
        from sqlalchemy.orm import joinedload
        contracts = db.query(PersonnelContract).options(
            joinedload(PersonnelContract.draft_contract)
        ).filter(
            PersonnelContract.is_active == 1
        ).all()
        
        for contract in contracts:
            personnel = db.query(Personnel).filter(
                Personnel.id == contract.personnel_id
            ).first()
            
            if not personnel:
                continue
            
            cost_center = None
            if contract.cost_center_id:
                cost_center = db.query(CostCenter).filter(
                    CostCenter.id == contract.cost_center_id
                ).first()
            
            ws.cell(row=row, column=1, value=str(personnel.tc_kimlik_no) if personnel.tc_kimlik_no else '')
            ws.cell(row=row, column=2, value=str(personnel.ad) if personnel.ad else '')
            ws.cell(row=row, column=3, value=str(personnel.soyad) if personnel.soyad else '')
            ws.cell(row=row, column=4, value=str(contract.bolum) if contract.bolum else '')
            ws.cell(row=row, column=5, value=contract.ise_giris_tarihi.strftime('%Y-%m-%d') if contract.ise_giris_tarihi else '')
            ws.cell(row=row, column=6, value=contract.isten_cikis_tarihi.strftime('%Y-%m-%d') if contract.isten_cikis_tarihi else '')
            # Draft contract verisi al (varsa)
            draft = contract.draft_contract if hasattr(contract, 'draft_contract') else None
            
            ws.cell(row=row, column=7, value=safe_enum_value(draft.ucret_nevi if draft else None, 'aylik'))
            ws.cell(row=row, column=8, value=safe_enum_value(contract.kanun_tipi, '4857'))
            ws.cell(row=row, column=9, value=safe_enum_value(draft.calisma_takvimi if draft else None, 'atipi'))
            ws.cell(row=row, column=10, value=safe_enum_value(contract.net_brut, 'B'))
            ws.cell(row=row, column=11, value=float(contract.ucret) if contract.ucret else 0)
            ws.cell(row=row, column=12, value=float(draft.net_ucret) if (draft and draft.net_ucret) else 0)
            ws.cell(row=row, column=13, value=safe_enum_value(draft.maas_hesabi if draft else None, 'tipa'))
            ws.cell(row=row, column=14, value=str(personnel.iban) if personnel.iban else '')
            ws.cell(row=row, column=15, value=float(draft.fm_orani) if (draft and draft.fm_orani) else 1.0)
            ws.cell(row=row, column=16, value=float(draft.tatil_orani) if (draft and draft.tatil_orani) else 1.0)
            ws.cell(row=row, column=17, value='Evet' if contract.taseron else 'Hayır')
            ws.cell(row=row, column=18, value=int(contract.taseron_id) if contract.taseron_id else '')
            ws.cell(row=row, column=19, value=safe_enum_value(contract.departman, ''))
            ws.cell(row=row, column=20, value=str(cost_center.code) if cost_center else '')
            row += 1
    
    # Kolon genişlikleri
    widths = [15, 15, 15, 30, 15, 15, 15, 12, 18, 12, 12, 12, 15, 30, 10, 10, 10, 12, 20, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else f"A{chr(64+i-26)}"].width = width
    
    # Excel'i byte olarak kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    donem_suffix = f"_{donem}" if donem else ""
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Personel_Sozlesmeler{donem_suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@router.post("/upload-excel")
async def upload_contracts_excel(
    file: UploadFile = File(...),
    current_user: UserInDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload personnel contracts from Excel file
    
    Args:
        file: Excel file (.xlsx or .xls)
    
    Returns:
        Upload results with counts and errors
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) yüklenebilir")
    
    try:
        # Excel'i oku
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        created_count = 0
        updated_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # TC kimlik kontrolü
                if pd.isna(row.get('TC Kimlik No')):
                    errors.append(f"Satır {idx+2}: TC Kimlik No boş")
                    continue
                
                tc = str(row['TC Kimlik No']).strip()
                
                # Personnel kontrolü
                personnel = db.query(Personnel).filter(Personnel.tc_kimlik_no == tc).first()
                if not personnel:
                    errors.append(f"Satır {idx+2}: TC {tc} için personel bulunamadı")
                    continue
                
                # Bölüm ve İşe Giriş Tarihi kontrolü
                if pd.isna(row.get('Bölüm')) or pd.isna(row.get('İşe Giriş Tarihi')):
                    errors.append(f"Satır {idx+2}: Bölüm veya İşe Giriş Tarihi boş")
                    continue
                
                bolum = str(row['Bölüm']).strip()
                giris_tarihi = pd.to_datetime(row['İşe Giriş Tarihi']).date()
                
                # Mevcut kayıt var mı kontrol
                existing = db.query(PersonnelContract).filter(
                    PersonnelContract.tc_kimlik_no == tc,
                    PersonnelContract.ise_giris_tarihi == giris_tarihi,
                    PersonnelContract.bolum == bolum
                ).first()
                
                # İşten Çıkış Tarihi
                cikis_tarihi = pd.to_datetime(row['İşten Çıkış Tarihi']).date() if pd.notna(row.get('İşten Çıkış Tarihi')) else None
                
                # Masraf Merkezi Kodu'ndan ID bul
                cost_center_id = None
                if pd.notna(row.get('Masraf Merkezi Kodu')):
                    cost_center = db.query(CostCenter).filter(
                        CostCenter.code == str(row['Masraf Merkezi Kodu']).strip()
                    ).first()
                    if cost_center:
                        cost_center_id = cost_center.id
                
                # Taşeron
                taseron_val = str(row.get('Taşeron', 'Hayır')).strip().lower()
                taseron = 1 if taseron_val in ['evet', '1', 'yes', 'true'] else 0
                
                # Taşeron ID kontrol
                taseron_id = None
                if pd.notna(row.get('Taşeron ID')):
                    taseron_id_val = int(row['Taşeron ID'])
                    contact_exists = db.query(Contact).filter(Contact.id == taseron_id_val).first()
                    if contact_exists:
                        taseron_id = taseron_id_val
                    else:
                        errors.append(f"Satır {idx+2}: Taşeron ID {taseron_id_val} contacts tablosunda bulunamadı")
                        continue
                
                # IBAN bilgisi
                iban_value = str(row['IBAN']).strip() if pd.notna(row.get('IBAN')) else personnel.iban
                if pd.notna(row.get('IBAN')) and iban_value:
                    personnel.iban = iban_value
                
                data = {
                    'personnel_id': personnel.id,
                    'tc_kimlik_no': tc,
                    'bolum': bolum,
                    'ise_giris_tarihi': giris_tarihi,
                    'isten_cikis_tarihi': cikis_tarihi,
                    'ucret_nevi': str(row['Ücret Nevi']).strip() if pd.notna(row.get('Ücret Nevi')) else 'aylik',
                    'kanun_tipi': str(row['Kanun Tipi']).strip() if pd.notna(row.get('Kanun Tipi')) else '4857',
                    'calisma_takvimi': str(row['Çalışma Takvimi']).strip() if pd.notna(row.get('Çalışma Takvimi')) else 'atipi',
                    'net_brut': str(row['Maaş1 Tip']).strip() if pd.notna(row.get('Maaş1 Tip')) else 'B',
                    'ucret': float(row['Maaş1 Tutar']) if pd.notna(row.get('Maaş1 Tutar')) else 0,
                    'maas2_tutar': float(row['Maaş2 Tutar']) if pd.notna(row.get('Maaş2 Tutar')) else None,
                    'maas_hesabi': str(row['Maaş Hesabı']).strip() if pd.notna(row.get('Maaş Hesabı')) else 'tipa',
                    'iban': iban_value,
                    'fm_orani': float(row['FM Oranı']) if pd.notna(row.get('FM Oranı')) else 1.0,
                    'tatil_orani': float(row['Tatil Oranı']) if pd.notna(row.get('Tatil Oranı')) else 1.0,
                    'taseron': taseron,
                    'taseron_id': taseron_id,
                    'departman': str(row['Departman']).strip() if pd.notna(row.get('Departman')) else None,
                    'cost_center_id': cost_center_id,
                    'is_active': 1 if not cikis_tarihi else 0
                }
                
                if existing:
                    # Güncelle
                    for key, value in data.items():
                        if key != 'personnel_id' and value is not None:
                            setattr(existing, key, value)
                    updated_count += 1
                else:
                    # Yeni kayıt
                    contract = PersonnelContract(**data)
                    db.add(contract)
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"Satır {idx+2}: {str(e)}")
                continue
        
        db.commit()
        
        return {
            "success": True,
            "message": f"{created_count} yeni sözleşme oluşturuldu, {updated_count} sözleşme güncellendi",
            "created_count": created_count,
            "updated_count": updated_count,
            "errors": errors[:20] if errors else []
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Dosya işleme hatası: {str(e)}")
