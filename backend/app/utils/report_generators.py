# -*- coding: utf-8 -*-
"""
Report Generation Utilities
PDF ve Excel rapor oluşturma fonksiyonları
"""
import io
from typing import BinaryIO, Dict, Any


def format_turkish_number(value: float) -> str:
    """Türkiye formatında sayı formatlar: 1.234.567,89"""
    formatted = f"{value:,.2f}"
    # İngilizce format: 1,234,567.89
    # Türkçe format: 1.234.567,89
    formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
    return formatted


def generate_cari_pdf(report_data: Dict[str, Any]) -> bytes:
    """Cari raporu için PDF oluşturur"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime
    
    # Türkçe karakter desteği
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        font_name = 'Arial'
        font_bold = 'Arial-Bold'
    except:
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Başlık
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#c41d7f'),
        alignment=1,
        fontName=font_bold
    )
    elements.append(Paragraph('CARİ HESAP EKSTRESİ', title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Tarih dönüşümü
    start_date = report_data['start_date'] if isinstance(report_data['start_date'], datetime) else datetime.fromisoformat(str(report_data['start_date']))
    end_date = report_data['end_date'] if isinstance(report_data['end_date'], datetime) else datetime.fromisoformat(str(report_data['end_date']))
    
    # Bilgi tablosu
    info_data = [
        ['Cari Adı:', report_data['contact_name'], '', 'Açılış Bakiyesi:', f"{format_turkish_number(float(report_data['opening_balance']))} TL"],
        ['Cari Kodu:', report_data['contact_code'] or '-', '', 'Toplam Borç:', f"{format_turkish_number(float(report_data['total_debit']))} TL"],
        ['Dönem:', f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}", '', 'Toplam Alacak:', f"{format_turkish_number(float(report_data['total_credit']))} TL"],
        ['', '', '', 'Kapanış Bakiyesi:', f"{format_turkish_number(float(report_data['closing_balance']))} TL"],
    ]
    
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 1*cm, 4*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),  # Tüm hücrelere font uygula
        ('FONTNAME', (0, 0), (0, -1), font_bold),
        ('FONTNAME', (3, 0), (3, -1), font_bold),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Hareket tablosu
    table_data = [['Tarih', 'Fiş No', 'Açıklama', 'Borç', 'Alacak', 'B/A', 'Bakiye']]
    
    # Açılış bakiyesi satırı ekle
    if report_data['opening_balance'] != 0:
        table_data.append([
            start_date.strftime('%d.%m.%Y'),
            '-',
            'AÇILIŞ BAKİYESİ',
            '',
            '',
            'B' if report_data['opening_balance'] > 0 else 'A' if report_data['opening_balance'] < 0 else '',
            format_turkish_number(abs(float(report_data['opening_balance'])))
        ])
    
    for item in report_data['items']:
        # Tarih dönüşümü
        trans_date = item['transaction_date'] if isinstance(item['transaction_date'], datetime) else datetime.fromisoformat(str(item['transaction_date']))
        
        table_data.append([
            trans_date.strftime('%d.%m.%Y'),
            item['transaction_number'],
            (item['description'] or '')[:50],
            format_turkish_number(float(item['debit'])),
            format_turkish_number(float(item['credit'])),
            'B' if item['balance'] > 0 else 'A' if item['balance'] < 0 else '',
            format_turkish_number(abs(float(item['balance'])))
        ])
    
    t = Table(table_data, colWidths=[2.5*cm, 3*cm, 8*cm, 3*cm, 3*cm, 1.5*cm, 3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(t)
    
    doc.build(elements)
    return output.getvalue()


def generate_cari_excel(report_data: Dict[str, Any]) -> bytes:
    """Cari raporu için Excel oluşturur"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cari Ekstre"
    
    # Başlık
    ws.merge_cells('A1:G1')
    ws['A1'] = 'CARİ HESAP EKSTRESİ'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Cari bilgileri
    ws['A3'] = 'Cari Adı:'
    ws['B3'] = report_data['contact_name']
    ws['A4'] = 'Cari Kodu:'
    ws['B4'] = report_data.get('contact_code') or '-'
    ws['A5'] = 'Dönem:'
    ws['B5'] = f"{report_data['start_date'].strftime('%d.%m.%Y')} - {report_data['end_date'].strftime('%d.%m.%Y')}"
    
    # Özet bilgiler (Türkçe sayı formatı ile)
    turkish_number_format = '#,##0.00;-#,##0.00'
    
    ws['D3'] = 'Açılış Bakiyesi:'
    ws['E3'] = float(report_data['opening_balance'])
    ws['E3'].number_format = turkish_number_format
    
    ws['D4'] = 'Toplam Borç:'
    ws['E4'] = float(report_data['total_debit'])
    ws['E4'].number_format = turkish_number_format
    
    ws['D5'] = 'Toplam Alacak:'
    ws['E5'] = float(report_data['total_credit'])
    ws['E5'].number_format = turkish_number_format
    
    ws['D6'] = 'Kapanış Bakiyesi:'
    ws['E6'] = float(report_data['closing_balance'])
    ws['E6'].number_format = turkish_number_format
    
    # Tablo başlıkları
    headers = ['Tarih', 'Fiş No', 'Açıklama', 'Borç', 'Alacak', 'B/A', 'Bakiye']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Veriler
    row = 9
    
    # Açılış bakiyesi satırı ekle
    opening_balance = report_data['opening_balance']
    if opening_balance != 0:
        ws.cell(row=row, column=1, value=report_data['start_date'].strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value='-')
        ws.cell(row=row, column=3, value='AÇILIŞ BAKİYESİ')
        ws.cell(row=row, column=4, value='')
        ws.cell(row=row, column=5, value='')
        ws.cell(row=row, column=6, value='B' if opening_balance > 0 else 'A' if opening_balance < 0 else '')
        
        opening_balance_cell = ws.cell(row=row, column=7, value=abs(float(opening_balance)))
        opening_balance_cell.number_format = turkish_number_format
        
        row += 1
    
    for item in report_data['items']:
        ws.cell(row=row, column=1, value=item['transaction_date'].strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=item['transaction_number'])
        ws.cell(row=row, column=3, value=item.get('description') or '')
        
        debit_cell = ws.cell(row=row, column=4, value=float(item['debit']))
        debit_cell.number_format = turkish_number_format
        
        credit_cell = ws.cell(row=row, column=5, value=float(item['credit']))
        credit_cell.number_format = turkish_number_format
        
        balance = item['balance']
        ws.cell(row=row, column=6, value='B' if balance > 0 else 'A' if balance < 0 else '')
        
        balance_cell = ws.cell(row=row, column=7, value=abs(float(balance)))
        balance_cell.number_format = turkish_number_format
        
        row += 1
    
    # Genişlikleri ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    
    # Excel dosyasını oluştur
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
