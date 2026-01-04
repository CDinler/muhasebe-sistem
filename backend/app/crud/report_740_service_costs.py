from sqlalchemy.orm import Session
from sqlalchemy import text
import pandas as pd
from datetime import date
from typing import Optional
from openpyxl.styles import Font, PatternFill, Alignment

def get_740_service_production_costs_excel(
    db: Session,
    start_date: date,
    end_date: date,
    cost_center_id: Optional[int] = None
) -> bytes:
    """
    740 Hizmet ve Üretim Maliyetleri raporu (Excel) - Detaylı versiyon
    Her bir harcama kalemi için evrak no, tarih, contact, açıklama ve tutar bilgileri
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    params = {'start_date': start_date, 'end_date': end_date}
    cost_center_filter = ''
    if cost_center_id:
        cost_center_filter = 'AND cc.id = :cost_center_id'
        params['cost_center_id'] = cost_center_id
    
    # Detaylı sorgu - her satır bir işlem hattı
    # Contact_id'ler güncellendi (120/320 cariler) - zaten doldurulan 20,900 satır var
    # 335 personeller için aynı fişten personnel bilgisini çek (subquery ile)
    sql = text(f"""
        SELECT 
            t.transaction_date as tarih,
            t.document_number as evrak_no,
            cc.code as cost_center_code,
            cc.name as cost_center_name,
            a.code as account_code,
            a.name as account_name,
            COALESCE(
                c.tax_number,
                (SELECT p.tckn 
                 FROM transaction_lines tl2
                 JOIN accounts a2 ON tl2.account_id = a2.id
                 JOIN personnel p ON a2.id = p.account_id
                 WHERE tl2.transaction_id = t.id
                   AND a2.code LIKE '335%'
                 LIMIT 1)
            ) as contact_code,
            COALESCE(
                c.name,
                (SELECT CONCAT(p.first_name, ' ', p.last_name)
                 FROM transaction_lines tl2
                 JOIN accounts a2 ON tl2.account_id = a2.id
                 JOIN personnel p ON a2.id = p.account_id
                 WHERE tl2.transaction_id = t.id
                   AND a2.code LIKE '335%'
                 LIMIT 1)
            ) as contact_name,
            tl.description as aciklama,
            tl.debit as borc,
            tl.credit as alacak,
            t.id as transaction_id,
            tl.id as line_id
        FROM transaction_lines tl
        JOIN accounts a ON tl.account_id = a.id
        JOIN transactions t ON tl.transaction_id = t.id
        JOIN cost_centers cc ON t.cost_center_id = cc.id
        LEFT JOIN contacts c ON tl.contact_id = c.id
        WHERE a.code LIKE '740%'
          AND t.transaction_date >= :start_date
          AND t.transaction_date <= :end_date
          {cost_center_filter}
        ORDER BY t.transaction_date, t.document_number, cc.name, a.code
    """)
    
    rows = db.execute(sql, params).fetchall()
    
    # Excel workbook oluştur
    wb = Workbook()
    
    # 1. Sayfa: Detaylı Hareketler
    ws_detail = wb.active
    ws_detail.title = "Detaylı Hareketler"
    
    detail_data = []
    for r in rows:
        detail_data.append({
            'Tarih': r.tarih,
            'Evrak No': r.evrak_no,
            'Maliyet Merkezi': f"{r.cost_center_code} - {r.cost_center_name}",
            'Hesap Kodu': r.account_code,
            'Hesap Adı': r.account_name,
            'Cari Kodu': r.contact_code or '',
            'Cari Adı': r.contact_name or '',
            'Açıklama': r.aciklama or '',
            'Borç': float(r.borc or 0),
            'Alacak': float(r.alacak or 0)
        })
    
    df_detail = pd.DataFrame(detail_data)
    
    # Detaylı sayfaya başlık ekle
    headers = list(df_detail.columns)
    ws_detail.append(headers)
    
    # Başlık satırını formatla
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, cell in enumerate(ws_detail[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Veriyi ekle
    for r_idx, row in enumerate(dataframe_to_rows(df_detail, index=False, header=False), 2):
        ws_detail.append(row)
    
    # Sütun genişliklerini ayarla
    ws_detail.column_dimensions['A'].width = 12  # Tarih
    ws_detail.column_dimensions['B'].width = 15  # Evrak No
    ws_detail.column_dimensions['C'].width = 25  # Maliyet Merkezi
    ws_detail.column_dimensions['D'].width = 12  # Hesap Kodu
    ws_detail.column_dimensions['E'].width = 35  # Hesap Adı
    ws_detail.column_dimensions['F'].width = 12  # Cari Kodu
    ws_detail.column_dimensions['G'].width = 30  # Cari Adı
    ws_detail.column_dimensions['H'].width = 40  # Açıklama
    ws_detail.column_dimensions['I'].width = 15  # Borç
    ws_detail.column_dimensions['J'].width = 15  # Alacak
    
    # 2. Sayfa: Özet (Maliyet Merkezi ve Hesap Bazında)
    ws_summary = wb.create_sheet("Özet")
    
    summary_sql = text(f"""
        SELECT 
            cc.code as cost_center_code,
            cc.name as cost_center_name,
            a.code as account_code,
            a.name as account_name,
            SUM(tl.debit) as total_debit,
            SUM(tl.credit) as total_credit,
            COUNT(DISTINCT t.id) as transaction_count
        FROM transaction_lines tl
        JOIN accounts a ON tl.account_id = a.id
        JOIN transactions t ON tl.transaction_id = t.id
        JOIN cost_centers cc ON t.cost_center_id = cc.id
        WHERE a.code LIKE '740%'
          AND t.transaction_date >= :start_date
          AND t.transaction_date <= :end_date
          {cost_center_filter}
        GROUP BY cc.code, cc.name, a.code, a.name
        ORDER BY cc.code, a.code
    """)
    
    summary_rows = db.execute(summary_sql, params).fetchall()
    
    summary_data = []
    for r in summary_rows:
        summary_data.append({
            'Maliyet Merkezi Kodu': r.cost_center_code,
            'Maliyet Merkezi': r.cost_center_name,
            'Hesap Kodu': r.account_code,
            'Hesap Adı': r.account_name,
            'İşlem Sayısı': r.transaction_count,
            'Toplam Borç': float(r.total_debit or 0),
            'Toplam Alacak': float(r.total_credit or 0),
            'Net (Borç-Alacak)': float(r.total_debit or 0) - float(r.total_credit or 0)
        })
    
    df_summary = pd.DataFrame(summary_data)
    
    # Özet sayfaya başlık ekle
    summary_headers = list(df_summary.columns)
    ws_summary.append(summary_headers)
    
    # Başlık satırını formatla
    for col_num, cell in enumerate(ws_summary[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Özet veriyi ekle
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=False), 2):
        ws_summary.append(row)
    
    # Özet sütun genişliklerini ayarla
    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 35
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 15
    ws_summary.column_dimensions['G'].width = 15
    ws_summary.column_dimensions['H'].width = 15
    
    # Excel'i kaydet
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()
