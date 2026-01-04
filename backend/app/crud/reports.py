# -*- coding: utf-8 -*-
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from datetime import date
from typing import Optional
from decimal import Decimal

from app.models.transaction import Transaction
from app.models.transaction_line import TransactionLine
from app.models.account import Account
from app.models.contact import Contact
from app.models.einvoice import EInvoice
from app.models.cost_center import CostCenter
from app.schemas.reports import (
    MizanReport, MizanItem,
    IncomeStatement, IncomeStatementItem,
    DebtorCreditorReport, DebtorCreditorItem,
    CariReport, CariReportItem,
    MuavinReport, MuavinItem
)


def get_mizan_report(
    db: Session,
    start_date: date,
    end_date: date
) -> MizanReport:
    """Mizan (Trial Balance) raporu oluşturur."""
    
    # Dönem başı bakiyeleri (start_date öncesi)
    opening_balances = db.query(
        Account.code,
        Account.name,
        func.sum(
            case(
                (TransactionLine.debit > 0, TransactionLine.debit),
                else_=0
            )
        ).label('opening_debit'),
        func.sum(
            case(
                (TransactionLine.credit > 0, TransactionLine.credit),
                else_=0
            )
        ).label('opening_credit')
    ).join(
        TransactionLine, Account.id == TransactionLine.account_id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date < start_date
    ).group_by(
        Account.id, Account.code, Account.name
    ).all()
    
    # Dönem hareketleri
    period_movements = db.query(
        Account.code,
        Account.name,
        func.sum(
            case(
                (TransactionLine.debit > 0, TransactionLine.debit),
                else_=0
            )
        ).label('period_debit'),
        func.sum(
            case(
                (TransactionLine.credit > 0, TransactionLine.credit),
                else_=0
            )
        ).label('period_credit')
    ).join(
        TransactionLine, Account.id == TransactionLine.account_id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by(
        Account.id, Account.code, Account.name
    ).all()
    
    # Verileri birleştir
    accounts_data = {}
    
    for row in opening_balances:
        accounts_data[row.code] = {
            'code': row.code,
            'name': row.name,
            'opening_debit': Decimal(str(row.opening_debit or 0)),
            'opening_credit': Decimal(str(row.opening_credit or 0)),
            'period_debit': Decimal('0'),
            'period_credit': Decimal('0')
        }
    
    for row in period_movements:
        if row.code not in accounts_data:
            accounts_data[row.code] = {
                'code': row.code,
                'name': row.name,
                'opening_debit': Decimal('0'),
                'opening_credit': Decimal('0'),
                'period_debit': Decimal('0'),
                'period_credit': Decimal('0')
            }
        accounts_data[row.code]['period_debit'] = Decimal(str(row.period_debit or 0))
        accounts_data[row.code]['period_credit'] = Decimal(str(row.period_credit or 0))
    
    # Mizan kalemlerini oluştur
    items = []
    totals = {
        'opening_debit': Decimal('0'),
        'opening_credit': Decimal('0'),
        'period_debit': Decimal('0'),
        'period_credit': Decimal('0'),
        'closing_debit': Decimal('0'),
        'closing_credit': Decimal('0')
    }
    
    for code in sorted(accounts_data.keys()):
        data = accounts_data[code]
        
        # Kapanış bakiyelerini hesapla
        closing_debit = data['opening_debit'] + data['period_debit']
        closing_credit = data['opening_credit'] + data['period_credit']
        
        item = MizanItem(
            account_code=data['code'],
            account_name=data['name'],
            opening_debit=data['opening_debit'],
            opening_credit=data['opening_credit'],
            period_debit=data['period_debit'],
            period_credit=data['period_credit'],
            closing_debit=closing_debit,
            closing_credit=closing_credit
        )
        items.append(item)
        
        # Toplamları güncelle
        totals['opening_debit'] += data['opening_debit']
        totals['opening_credit'] += data['opening_credit']
        totals['period_debit'] += data['period_debit']
        totals['period_credit'] += data['period_credit']
        totals['closing_debit'] += closing_debit
        totals['closing_credit'] += closing_credit
    
    return MizanReport(
        start_date=start_date,
        end_date=end_date,
        items=items,
        **totals
    )


def get_income_statement(
    db: Session,
    start_date: date,
    end_date: date
) -> IncomeStatement:
    """Gelir-Gider tablosu oluşturur."""
    
    # Gelir ve gider hesaplarını getir (6xx gelir, 7xx gider)
    results = db.query(
        Account.code,
        Account.name,
        func.sum(TransactionLine.credit - TransactionLine.debit).label('amount')
    ).join(
        TransactionLine, Account.id == TransactionLine.account_id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
        Account.code.like('6%') | Account.code.like('7%')
    ).group_by(
        Account.id, Account.code, Account.name
    ).all()
    
    income_items = []
    expense_items = []
    total_income = Decimal('0')
    total_expense = Decimal('0')
    
    for row in results:
        amount = Decimal(str(row.amount or 0))
        
        if row.code.startswith('6'):  # Gelir hesapları
            income_items.append(IncomeStatementItem(
                account_code=row.code,
                account_name=row.name,
                amount=amount
            ))
            total_income += amount
        elif row.code.startswith('7'):  # Gider hesapları
            expense_items.append(IncomeStatementItem(
                account_code=row.code,
                account_name=row.name,
                amount=abs(amount)
            ))
            total_expense += abs(amount)
    
    net_profit = total_income - total_expense
    
    return IncomeStatement(
        start_date=start_date,
        end_date=end_date,
        income_items=income_items,
        expense_items=expense_items,
        total_income=total_income,
        total_expense=total_expense,
        net_profit=net_profit
    )


def get_debtor_creditor_report(
    db: Session,
    start_date: date,
    end_date: date
) -> DebtorCreditorReport:
    """Borçlu-Alacaklı raporu oluşturur."""
    
    results = db.query(
        Contact.id,
        Contact.name,
        Contact.tax_number,
        func.sum(TransactionLine.debit).label('total_debit'),
        func.sum(TransactionLine.credit).label('total_credit')
    ).join(
        TransactionLine, Contact.id == TransactionLine.contact_id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
        TransactionLine.contact_id.isnot(None)
    ).group_by(
        Contact.id, Contact.name, Contact.tax_number
    ).all()
    
    debtors = []
    creditors = []
    total_debtors = Decimal('0')
    total_creditors = Decimal('0')
    
    for row in results:
        debit = Decimal(str(row.total_debit or 0))
        credit = Decimal(str(row.total_credit or 0))
        balance = debit - credit
        
        item = DebtorCreditorItem(
            contact_id=row.id,
            contact_name=row.name,
            tax_number=row.tax_number,
            debit=debit,
            credit=credit,
            balance=balance
        )
        
        if balance > 0:  # Borçlu
            debtors.append(item)
            total_debtors += balance
        elif balance < 0:  # Alacaklı
            creditors.append(item)
            total_creditors += abs(balance)
    
    net_balance = total_debtors - total_creditors
    
    return DebtorCreditorReport(
        start_date=start_date,
        end_date=end_date,
        debtors=debtors,
        creditors=creditors,
        total_debtors=total_debtors,
        total_creditors=total_creditors,
        net_balance=net_balance
    )


def get_cari_report(
    db: Session,
    start_date: date,
    end_date: date,
    contact_id: Optional[int] = None
) -> CariReport:
    """
    Cari Raporu - Fiş fiş detaylı rapor
    120.xxx = Alıcılar (Müşteriler)
    320.xxx = Satıcılar (Tedarikçiler)
    
    contact_id verilirse o cariye ait hesapları döndürür.
    Eğer aynı VKN/isimde hem 120 hem 320 hesabı varsa ikisini de birleştirir.
    """
    
    # İlgili contact bilgilerini al ve ilişkili tüm contact'ları bul
    contact_info = None
    related_account_codes = []
    
    if contact_id:
        contact = db.query(Contact).filter(Contact.id == contact_id).first()
        if contact:
            contact_info = {
                'contact_id': contact.id,
                'contact_code': contact.code,
                'contact_name': contact.name,
                'tax_number': contact.tax_number
            }
            
            # Contact'ın hesap kodundan sayı kısmını al (örn: 320.00547 -> 00547)
            if contact.code:
                code_parts = contact.code.split('.')
                if len(code_parts) == 2:
                    number_part = code_parts[1]  # 00547
                    
                    # Hem 120.XXXXX hem 320.XXXXX hesaplarını ekle
                    related_account_codes = [
                        f"120.{number_part}",
                        f"320.{number_part}"
                    ]
                else:
                    related_account_codes = [contact.code]
            else:
                related_account_codes = []
    
    # Dönem başı bakiyesi (start_date öncesi)
    opening_query = db.query(
        func.sum(
            case((TransactionLine.debit > 0, TransactionLine.debit), else_=0)
        ).label('opening_debit'),
        func.sum(
            case((TransactionLine.credit > 0, TransactionLine.credit), else_=0)
        ).label('opening_credit')
    ).select_from(TransactionLine
    ).join(
        Account, TransactionLine.account_id == Account.id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date < start_date,
        (Account.code.like('120.%') | Account.code.like('320.%'))
    )
    
    if related_account_codes:
        # İlişkili tüm hesap kodlarını filtrele
        opening_query = opening_query.filter(Account.code.in_(related_account_codes))
    
    opening_result = opening_query.first()
    opening_debit = Decimal(str(opening_result.opening_debit or 0)) if opening_result else Decimal('0')
    opening_credit = Decimal(str(opening_result.opening_credit or 0)) if opening_result else Decimal('0')
    opening_balance = opening_debit - opening_credit
    
    # Dönem içi hareketler (FİŞ FİŞ - Her fiş için TOPLAM borç/alacak VE hesap kodu)
    movements_query = db.query(
        Transaction.id.label('transaction_id'),
        Transaction.transaction_number,
        Transaction.transaction_date,
        Transaction.description,
        func.group_concat(func.distinct(Account.code)).label('account_codes'),  # Fişdeki tüm 120/320 kodları
        func.sum(TransactionLine.debit).label('total_debit'),
        func.sum(TransactionLine.credit).label('total_credit')
    ).select_from(TransactionLine
    ).join(
        Account, TransactionLine.account_id == Account.id
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
        (Account.code.like('120.%') | Account.code.like('320.%'))
    )
    
    if related_account_codes:
        # İlişkili tüm hesap kodlarını filtrele
        movements_query = movements_query.filter(Account.code.in_(related_account_codes))
    
    movements_query = movements_query.group_by(
        Transaction.id,
        Transaction.transaction_number,
        Transaction.transaction_date,
        Transaction.description
    ).order_by(
        Transaction.transaction_date,
        Transaction.transaction_number
    )
    
    movements = movements_query.all()
    
    # İlk satırdan contact bilgisi al
    if not contact_info and movements:
        # İlk fişin hesap kodunu bul
        first_tx = movements[0]
        first_account = db.query(Account).join(
            TransactionLine, Account.id == TransactionLine.account_id
        ).filter(
            TransactionLine.transaction_id == first_tx.transaction_id,
            (Account.code.like('120.%') | Account.code.like('320.%'))
        ).first()
        
        if first_account:
            first_contact = db.query(Contact).filter(Contact.code == first_account.code).first()
            if first_contact:
                contact_info = {
                    'contact_id': first_contact.id,
                    'contact_code': first_contact.code,
                    'contact_name': first_contact.name,
                    'tax_number': first_contact.tax_number
                }
            else:
                contact_info = {
                    'contact_id': None,
                    'contact_code': first_account.code,
                    'contact_name': first_account.name,
                    'tax_number': None
                }
    
    if not contact_info:
        contact_info = {
            'contact_id': None,
            'contact_code': None,
            'contact_name': 'Genel',
            'tax_number': None
        }
    
    # Fiş bazında rapor oluştur
    items = []
    running_balance = opening_balance
    total_debit = Decimal('0')
    total_credit = Decimal('0')
    
    for row in movements:
        debit = Decimal(str(row.total_debit or 0))
        credit = Decimal(str(row.total_credit or 0))
        running_balance = running_balance + debit - credit
        
        # Bu fişin hesap kodunu al (ilk kodu kullan)
        account_codes = row.account_codes.split(',') if row.account_codes else []
        account_code = account_codes[0] if account_codes else (contact_info['contact_code'] or '')
        
        # Hesap adını bul
        if account_code:
            account = db.query(Account).filter(Account.code == account_code).first()
            account_name = account.name if account else contact_info['contact_name']
        else:
            account_name = contact_info['contact_name']
        
        # Evrak türünü belirle
        document_type = None
        if row.description:
            if 'FATURA' in row.description.upper() or 'İHRACAT' in row.description.upper():
                document_type = 'İhracat Faturası'
            elif 'HAVALE' in row.description.upper() or 'GELEN' in row.description.upper():
                document_type = 'Gelen Havale'
            elif 'İRSALIYE' in row.description.upper():
                document_type = 'İrsaliye'
            else:
                document_type = row.description[:50]
        
        # Hesap tipi
        account_type = 'customer' if account_code.startswith('120') else 'supplier'
        
        item = CariReportItem(
            transaction_id=row.transaction_id,
            transaction_number=row.transaction_number,
            transaction_date=row.transaction_date,
            due_date=None,
            document_type=document_type,
            description=row.description,
            account_code=account_code,
            account_name=account_name,
            account_type=account_type,
            currency='TL',
            currency_debit=debit,
            currency_credit=credit,
            currency_balance=running_balance,
            debit=debit,
            credit=credit,
            balance=running_balance
        )
        
        items.append(item)
        total_debit += debit
        total_credit += credit
    
    return CariReport(
        contact_id=contact_info['contact_id'],
        contact_code=contact_info['contact_code'],
        contact_name=contact_info['contact_name'],
        tax_number=contact_info['tax_number'],
        start_date=start_date,
        end_date=end_date,
        opening_balance=opening_balance,
        items=items,
        closing_balance=running_balance,
        total_debit=total_debit,
        total_credit=total_credit
    )


def get_muavin_report(
    db: Session,
    account_code: str,
    start_date: date,
    end_date: date
) -> MuavinReport:
    """
    Muavin Defteri - Belirli bir hesap kodunun tüm işlemleri
    
    Args:
        account_code: Hesap kodu (örn: "100", "120.00001")
        start_date: Başlangıç tarihi
        end_date: Bitiş tarihi
    """
    
    # Hesap bilgisini al
    account = db.query(Account).filter(Account.code == account_code).first()
    if not account:
        # Hesap bulunamazsa boş rapor döndür
        return MuavinReport(
            account_code=account_code,
            account_name="Hesap bulunamadı",
            start_date=start_date,
            end_date=end_date,
            opening_balance=Decimal('0'),
            items=[],
            closing_balance=Decimal('0'),
            total_debit=Decimal('0'),
            total_credit=Decimal('0')
        )
    
    # Dönem başı bakiyesi (start_date öncesi)
    opening_query = db.query(
        func.sum(
            case((TransactionLine.debit > 0, TransactionLine.debit), else_=0)
        ).label('opening_debit'),
        func.sum(
            case((TransactionLine.credit > 0, TransactionLine.credit), else_=0)
        ).label('opening_credit')
    ).select_from(TransactionLine
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        TransactionLine.account_id == account.id,
        Transaction.transaction_date < start_date
    )
    
    opening_result = opening_query.first()
    opening_debit = Decimal(str(opening_result.opening_debit or 0)) if opening_result else Decimal('0')
    opening_credit = Decimal(str(opening_result.opening_credit or 0)) if opening_result else Decimal('0')
    opening_balance = opening_debit - opening_credit
    
    # Dönem içi hareketler
    movements = db.query(
        Transaction.id.label('transaction_id'),
        Transaction.transaction_number,
        Transaction.transaction_date,
        Transaction.description,
        TransactionLine.debit,
        TransactionLine.credit
    ).select_from(TransactionLine
    ).join(
        Transaction, TransactionLine.transaction_id == Transaction.id
    ).filter(
        TransactionLine.account_id == account.id,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).order_by(
        Transaction.transaction_date,
        Transaction.transaction_number
    ).all()
    
    # İşlemleri oluştur
    items = []
    running_balance = opening_balance
    total_debit = Decimal('0')
    total_credit = Decimal('0')
    
    for row in movements:
        debit = Decimal(str(row.debit or 0))
        credit = Decimal(str(row.credit or 0))
        running_balance = running_balance + debit - credit
        
        item = MuavinItem(
            transaction_id=row.transaction_id,
            transaction_number=row.transaction_number,
            transaction_date=row.transaction_date,
            description=row.description,
            debit=debit,
            credit=credit,
            balance=running_balance
        )
        
        items.append(item)
        total_debit += debit
        total_credit += credit
    
    return MuavinReport(
        account_code=account.code,
        account_name=account.name,
        start_date=start_date,
        end_date=end_date,
        opening_balance=opening_balance,
        items=items,
        closing_balance=running_balance,
        total_debit=total_debit,
        total_credit=total_credit
    )


def get_yevmiye_excel(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    """Yevmiye Defteri Excel raporu oluşturur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from app.models.document_type import DocumentType, DocumentSubtype
    
    # Sorgu oluştur
    query = db.query(
        Transaction.transaction_date,
        Transaction.transaction_number,
        Transaction.accounting_period,
        Transaction.document_number,
        DocumentType.name.label('document_type'),
        DocumentSubtype.name.label('document_subtype'),
        Transaction.description,
        TransactionLine.id.label('line_id'),
        Account.code.label('account_code'),
        Account.name.label('account_name'),
        TransactionLine.description.label('line_description'),
        TransactionLine.debit,
        TransactionLine.credit,
        CostCenter.name.label('cost_center_name'),
        Transaction.related_invoice_number,
        EInvoice.invoice_number.label('einvoice_number')
    ).join(
        TransactionLine, Transaction.id == TransactionLine.transaction_id
    ).join(
        Account, TransactionLine.account_id == Account.id
    ).outerjoin(
        CostCenter, Transaction.cost_center_id == CostCenter.id
    ).outerjoin(
        DocumentType, Transaction.document_type_id == DocumentType.id
    ).outerjoin(
        DocumentSubtype, Transaction.document_subtype_id == DocumentSubtype.id
    ).outerjoin(
        EInvoice, Transaction.id == EInvoice.transaction_id
    )
    
    # Tarih filtresi
    if start_date:
        query = query.filter(Transaction.transaction_date >= start_date)
    if end_date:
        query = query.filter(Transaction.transaction_date <= end_date)
    
    # Sıralama: Tarih ve fiş numarasına göre
    query = query.order_by(
        Transaction.transaction_date.asc(),
        Transaction.transaction_number.asc(),
        TransactionLine.id.asc()
    )
    
    records = query.all()
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Yevmiye Defteri"
    
    # Başlık stili
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        'Tarih', 'Fiş No', 'Dönem', 'Evrak No', 'Evrak Türü', 'Evrak Alt Türü',
        'Açıklama', 'Satır', 'Hesap Kodu', 'Hesap Adı', 'Satır Açıklama',
        'Borç', 'Alacak', 'Masraf Merkezi', 'İlgili Fatura No'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Veri satırları
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.transaction_date.strftime('%d.%m.%Y') if record.transaction_date else '')
        ws.cell(row=row_idx, column=2, value=record.transaction_number or '')
        ws.cell(row=row_idx, column=3, value=record.accounting_period or '')
        ws.cell(row=row_idx, column=4, value=record.document_number or '')
        ws.cell(row=row_idx, column=5, value=record.document_type or '')
        ws.cell(row=row_idx, column=6, value=record.document_subtype or '')
        ws.cell(row=row_idx, column=7, value=record.description or '')
        ws.cell(row=row_idx, column=8, value=record.line_id or '')
        ws.cell(row=row_idx, column=9, value=record.account_code or '')
        ws.cell(row=row_idx, column=10, value=record.account_name or '')
        ws.cell(row=row_idx, column=11, value=record.line_description or '')
        
        # Sayısal değerler
        debit_cell = ws.cell(row=row_idx, column=12, value=float(record.debit) if record.debit else 0)
        debit_cell.number_format = '#,##0.00'
        
        credit_cell = ws.cell(row=row_idx, column=13, value=float(record.credit) if record.credit else 0)
        credit_cell.number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=14, value=record.cost_center_name or '')
        
        # İlgili Fatura No - önce e-fatura numarasına bak, yoksa related_invoice_number kullan
        invoice_ref = record.einvoice_number or record.related_invoice_number or ''
        ws.cell(row=row_idx, column=15, value=invoice_ref)
        
        # Kenarlıklar
        for col in range(1, 16):
            ws.cell(row=row_idx, column=col).border = border
    
    # Sütun genişlikleri
    column_widths = [12, 15, 10, 15, 15, 15, 40, 8, 15, 35, 40, 15, 15, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Excel'i byte olarak döndür
    from io import BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def get_personnel_accounts_excel(db: Session) -> bytes:
    """
    335 Personel Hesap Raporu Excel - 2025-11 döneminde çalışanların hesapları ve bakiyeleri
    HIZLI VERSİYON: Bakiyeler tek sorguda hesaplanıyor
    """
    from app.models.personnel import Personnel
    from app.models.luca_bordro import LucaBordro
    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # ADIM 1: Tüm 335 hesaplarının bakiyelerini TEK SORGUDA hesapla
    balances_result = db.execute(text("""
        SELECT 
            a.code,
            SUM(tl.debit) as total_debit,
            SUM(tl.credit) as total_credit,
            SUM(tl.debit) - SUM(tl.credit) as balance
        FROM accounts a
        LEFT JOIN transaction_lines tl ON tl.account_id = a.id
        WHERE a.code LIKE '335.%'
        GROUP BY a.code
    """))
    
    # Bakiyeleri dictionary'e aktar (hızlı erişim için)
    balances_dict = {row.code: {
        'debit': float(row.total_debit or 0),
        'credit': float(row.total_credit or 0),
        'balance': float(row.balance or 0)
    } for row in balances_result}
    
    # ADIM 2: 2025-11 bordrosundaki personelleri getir
    query = db.query(
        Personnel.tckn,
        Personnel.first_name,
        Personnel.last_name,
        Personnel.department,
        Personnel.position,
        Personnel.is_active,
        Account.code.label('account_code'),
        Account.name.label('account_name')
    ).join(
        LucaBordro, Personnel.tckn == LucaBordro.tckn
    ).outerjoin(
        Account, Personnel.account_id == Account.id
    ).filter(
        LucaBordro.donem == '2025-11'
    ).distinct().order_by(
        Personnel.first_name,
        Personnel.last_name
    )
    
    personnel_list = query.all()
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "335 Personel (2025-11)"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ['TC', 'Ad Soyad', 'Departman', 'Pozisyon', 'Hesap Kodu', 'Hesap Adı', 'Borç', 'Alacak', 'Bakiye', 'Durum']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Veri satırları
    total_debit = 0.0
    total_credit = 0.0
    total_balance = 0.0
    active_count = 0
    
    for row_idx, p in enumerate(personnel_list, 2):
        # Dictionary'den bakiye al (çok hızlı!)
        balance_info = balances_dict.get(p.account_code, {'debit': 0, 'credit': 0, 'balance': 0})
        debit = balance_info['debit']
        credit = balance_info['credit']
        balance = balance_info['balance']
        
        total_debit += debit
        total_credit += credit
        total_balance += balance
        
        if p.is_active:
            active_count += 1
        
        # Satır verileri
        ws.cell(row=row_idx, column=1, value=p.tckn)
        ws.cell(row=row_idx, column=2, value=f"{p.first_name} {p.last_name}")
        ws.cell(row=row_idx, column=3, value=p.department or '')
        ws.cell(row=row_idx, column=4, value=p.position or '')
        ws.cell(row=row_idx, column=5, value=p.account_code or 'Hesap Yok')
        ws.cell(row=row_idx, column=6, value=p.account_name or '')
        
        # Borç
        debit_cell = ws.cell(row=row_idx, column=7, value=debit)
        debit_cell.number_format = '#,##0.00'
        
        # Alacak
        credit_cell = ws.cell(row=row_idx, column=8, value=credit)
        credit_cell.number_format = '#,##0.00'
        
        # Bakiye
        balance_cell = ws.cell(row=row_idx, column=9, value=balance)
        balance_cell.number_format = '#,##0.00'
        
        # Bakiye renklendirme
        if balance > 0:
            balance_cell.font = Font(color="CF1322", bold=True)
        elif balance < 0:
            balance_cell.font = Font(color="3F8600", bold=True)
        
        ws.cell(row=row_idx, column=10, value='Aktif' if p.is_active else 'Pasif')
        
        # Kenarlıklar
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = border
    
    # Özet satırları
    summary_row = len(personnel_list) + 3
    ws.cell(row=summary_row, column=1, value='ÖZET (2025-11 Dönemi)').font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value=f'Toplam Personel: {len(personnel_list)}')
    ws.cell(row=summary_row + 2, column=1, value=f'Aktif Personel: {active_count}')
    
    ws.cell(row=summary_row + 3, column=1, value='Toplam Borç:')
    ws.cell(row=summary_row + 3, column=2, value=total_debit).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 4, column=1, value='Toplam Alacak:')
    ws.cell(row=summary_row + 4, column=2, value=total_credit).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 5, column=1, value='Net Bakiye:')
    total_cell = ws.cell(row=summary_row + 5, column=2, value=total_balance)
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True, size=12)
    
    # Sütun genişlikleri
    column_widths = [15, 30, 20, 20, 18, 35, 15, 15, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Excel'i byte olarak döndür
    from io import BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
