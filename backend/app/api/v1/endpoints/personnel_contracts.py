"""
Personnel Contracts API endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.database import get_db
from app.models.personnel_contract import PersonnelContract
from app.models.personnel import Personnel
from app.models.cost_center import CostCenter
from app.schemas.personnel_contract import (
    PersonnelContractCreate,
    PersonnelContractUpdate,
    PersonnelContractResponse,
    PersonnelContractList
)

router = APIRouter()


@router.get("/template")
async def download_contracts_template(db: Session = Depends(get_db)):
    """
    Personel sözleşme şablonu indir
    Mevcut personel listesi + sözleşme bilgileri (varsa)
    """
    # Tüm personel listesi
    personnel_list = db.query(Personnel).all()
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel Sözleşmeleri"
    
    # Header
    headers = [
        'TC', 'Ad', 'Soyad', 'Başlangıç Tarihi', 'Bitiş Tarihi',
        'Bölüm', 'Taşeron', 'Departman', 'Pozisyon', 'Ünvan',
        'Maaş Hesabı', 'Çalışma Takvimi', 'Sigorta Durumu', 'IBAN'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Personel verilerini ekle (mevcut sözleşmeler varsa doldur)
    row = 2
    for person in personnel_list:
        # Son aktif sözleşmeyi bul
        contract = db.query(PersonnelContract).filter(
            PersonnelContract.personnel_id == person.id,
            PersonnelContract.aktif == True
        ).first()
        
        ws.cell(row=row, column=1, value=str(person.tc_kimlik_no) if person.tc_kimlik_no else '')
        ws.cell(row=row, column=2, value=str(person.ad) if person.ad else '')
        ws.cell(row=row, column=3, value=str(person.soyad) if person.soyad else '')
        ws.cell(row=row, column=4, value=contract.baslangic_tarihi.strftime('%Y-%m-%d') if (contract and contract.baslangic_tarihi) else '')
        ws.cell(row=row, column=5, value=contract.bitis_tarihi.strftime('%Y-%m-%d') if (contract and contract.bitis_tarihi) else '')
        ws.cell(row=row, column=6, value=str(contract.bolum) if (contract and contract.bolum) else '')
        ws.cell(row=row, column=7, value='Evet' if (contract and contract.taseron) else 'Hayır')
        ws.cell(row=row, column=8, value=str(contract.departman.value) if (contract and contract.departman) else '')
        ws.cell(row=row, column=9, value=str(contract.pozisyon) if (contract and contract.pozisyon) else '')
        ws.cell(row=row, column=10, value=str(contract.unvan) if (contract and contract.unvan) else '')
        ws.cell(row=row, column=11, value=str(contract.maas_hesabi.value) if (contract and contract.maas_hesabi) else '')
        ws.cell(row=row, column=12, value=str(contract.calisma_takvimi.value) if (contract and contract.calisma_takvimi) else '')
        ws.cell(row=row, column=13, value=str(contract.sigorta_durumu.value) if (contract and contract.sigorta_durumu) else '')
        ws.cell(row=row, column=14, value=str(person.iban) if person.iban else '')
        row += 1
    
    # Kolon genişlikleri
    for i, width in enumerate([12, 20, 20, 15, 15, 25, 10, 20, 20, 20, 15, 15, 15, 30], 1):
        ws.column_dimensions[chr(64+i)].width = width
    
    # Excel'i byte olarak kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Personel_Sozlesmeler_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@router.get("/list", response_model=PersonnelContractList)
def list_contracts(
    personnel_id: Optional[int] = None,
    aktif: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """
    Sözleşmeleri listele
    """
    query = db.query(PersonnelContract)
    
    if personnel_id:
        query = query.filter(PersonnelContract.personnel_id == personnel_id)
    
    if aktif is not None:
        query = query.filter(PersonnelContract.aktif == aktif)
    
    total = query.count()
    contracts = query.order_by(PersonnelContract.baslangic_tarihi.desc()).offset(skip).limit(limit).all()
    
    return PersonnelContractList(
        total=total,
        page=skip // limit + 1 if limit > 0 else 1,
        page_size=limit,
        items=[PersonnelContractResponse.model_validate(c) for c in contracts]
    )


@router.get("/{contract_id}", response_model=PersonnelContractResponse)
def get_contract(
    contract_id: int,
    db: Session = Depends(get_db)
):
    """
    Sözleşme detayını getir
    """
    contract = db.query(PersonnelContract).filter(PersonnelContract.id == contract_id).first()
    
    if not contract:
        raise HTTPException(status_code=404, detail="Sözleşme bulunamadı")
    
    return contract


@router.post("/create", response_model=PersonnelContractResponse)
def create_contract(
    data: PersonnelContractCreate,
    db: Session = Depends(get_db)
):
    """
    Yeni sözleşme oluştur
    """
    # Personnel var mı kontrol et
    personnel = db.query(Personnel).filter(Personnel.id == data.personnel_id).first()
    if not personnel:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    
    # Aynı tarihte aktif sözleşme var mı
    if data.baslangic_tarihi:
        existing = db.query(PersonnelContract).filter(
            PersonnelContract.personnel_id == data.personnel_id,
            PersonnelContract.baslangic_tarihi == data.baslangic_tarihi,
            PersonnelContract.aktif == True
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Bu tarihte aktif sözleşme zaten mevcut")
    
    contract = PersonnelContract(**data.model_dump())
    db.add(contract)
    db.commit()
    db.refresh(contract)
    
    return contract


@router.put("/{contract_id}", response_model=PersonnelContractResponse)
def update_contract(
    contract_id: int,
    data: PersonnelContractUpdate,
    db: Session = Depends(get_db)
):
    """
    Sözleşme güncelle
    """
    contract = db.query(PersonnelContract).filter(PersonnelContract.id == contract_id).first()
    
    if not contract:
        raise HTTPException(status_code=404, detail="Sözleşme bulunamadı")
    
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(contract, key, value)
    
    db.commit()
    db.refresh(contract)
    
    return contract


@router.delete("/{contract_id}")
def delete_contract(
    contract_id: int,
    db: Session = Depends(get_db)
):
    """
    Sözleşmeyi sil (soft delete - aktif=False)
    """
    contract = db.query(PersonnelContract).filter(PersonnelContract.id == contract_id).first()
    
    if not contract:
        raise HTTPException(status_code=404, detail="Sözleşme bulunamadı")
    
    contract.aktif = False
    db.commit()
    
    return {"success": True, "message": "Sözleşme pasif hale getirildi"}


@router.post("/upload-excel")
async def upload_contracts_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Sözleşmeleri Excel'den toplu yükle
    
    Beklenen sütunlar:
    TC Kimlik No | Başlangıç Tarihi | Bitiş Tarihi | Bölüm | Departman | Pozisyon | Ünvan
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        uploaded_count = 0
        updated_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # TC Kimlik No (zorunlu)
                tc_col = row.get('TC Kimlik No') or row.get('TC') or row.get('TCKN')
                tc = str(int(tc_col)) if pd.notna(tc_col) else None
                
                if not tc or len(tc) != 11:
                    errors.append(f"Satır {idx+2}: TC bulunamadı veya geçersiz")
                    continue
                
                # Personnel bul
                personnel = db.query(Personnel).filter(Personnel.tc_kimlik_no == tc).first()
                
                if not personnel:
                    errors.append(f"Satır {idx+2}: TC {tc} sistemde yok")
                    continue
                
                # Başlangıç Tarihi
                baslangic_col = row.get('Başlangıç Tarihi') or row.get('İşe Giriş Tarihi')
                baslangic = pd.to_datetime(baslangic_col).date() if pd.notna(baslangic_col) else None
                
                if not baslangic:
                    errors.append(f"Satır {idx+2}: Başlangıç tarihi bulunamadı")
                    continue
                
                # Aynı kayıt var mı
                existing = db.query(PersonnelContract).filter(
                    PersonnelContract.personnel_id == personnel.id,
                    PersonnelContract.baslangic_tarihi == baslangic
                ).first()
                
                # Bitiş Tarihi
                bitis_col = row.get('Bitiş Tarihi') or row.get('İşten Çıkış Tarihi')
                bitis = pd.to_datetime(bitis_col).date() if pd.notna(bitis_col) else None
                
                data = {
                    'personnel_id': personnel.id,
                    'tc_kimlik_no': tc,
                    'baslangic_tarihi': baslangic,
                    'bitis_tarihi': bitis,
                    'bolum': str(row['Bölüm']).strip() if pd.notna(row.get('Bölüm')) else None,
                    'pozisyon': str(row['Pozisyon']).strip() if pd.notna(row.get('Pozisyon')) else None,
                    'unvan': str(row['Ünvan']).strip() if pd.notna(row.get('Ünvan')) else None,
                    'aktif': True if not bitis else False
                }
                
                if existing:
                    for key, value in data.items():
                        if key != 'personnel_id':
                            setattr(existing, key, value)
                    updated_count += 1
                else:
                    contract = PersonnelContract(**data)
                    db.add(contract)
                    uploaded_count += 1
                    
            except Exception as e:
                errors.append(f"Satır {idx+2}: {str(e)}")
                continue
        
        db.commit()
        
        return {
            "success": True,
            "message": f"{uploaded_count} yeni sözleşme, {updated_count} güncelleme yapıldı",
            "uploaded_count": uploaded_count,
            "updated_count": updated_count,
            "errors": errors[:20] if errors else []
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Dosya işleme hatası: {str(e)}")

